from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple
import csv
import re
import uuid
import math
import os
import json
import sqlite3
import time
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
except Exception:  # pragma: no cover
    ZoneInfo = None  # Fallback if not available
import shutil

from flask import Flask, render_template, request, redirect, url_for, send_file, flash, session, jsonify
from openpyxl import Workbook, load_workbook
import google.generativeai as genai


app = Flask(__name__)
app.secret_key = "order-stock-secret-key"

# Database configuration
DATABASE_PATH = os.path.join(os.path.dirname(__file__), 'orderai.sqlite3')
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# In-memory result store keyed by a one-time token
RESULT_STORE: Dict[str, Dict[str, bytes]] = {}
# Batch store maps a batch token to a list of item tokens
BATCH_STORE: Dict[str, List[str]] = {}
# Simple schedule store: token -> ISO date string
SCHEDULE_STORE: Dict[str, str] = {}
# Decisions store for Not have items: key(fish|pack|order) -> decision
DECISION_STORE: Dict[str, str] = {}


# ----- Database Functions -----
def init_database():
    """Initialize the SQLite database with required tables."""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    # Create processing sessions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS processing_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_token TEXT UNIQUE NOT NULL,
            batch_token TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            processing_type TEXT NOT NULL, -- 'single' or 'batch'
            total_items INTEGER DEFAULT 0,
            full_items INTEGER DEFAULT 0,
            not_full_items INTEGER DEFAULT 0,
            not_have_items INTEGER DEFAULT 0,
            total_kg REAL DEFAULT 0,
            full_kg REAL DEFAULT 0,
            not_full_kg REAL DEFAULT 0,
            not_have_kg REAL DEFAULT 0
        )
    ''')
    
    # Create uploaded files table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS uploaded_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER,
            file_type TEXT NOT NULL, -- 'stock', 'import_stock', or 'order'
            original_filename TEXT NOT NULL,
            stored_filename TEXT NOT NULL,
            file_size INTEGER NOT NULL,
            is_revision BOOLEAN DEFAULT 0, -- 0 for original, 1 for revised
            revision_note TEXT, -- Note about the revision
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (session_id) REFERENCES processing_sessions (id)
        )
    ''')
    
    # Create processing results table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS processing_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER,
            order_file_id INTEGER,
            fish_name TEXT,
            packed_size TEXT,
            order_carton INTEGER,
            stock_carton INTEGER,
            order_kg_per_ctn REAL,
            stock_kg_per_ctn REAL,
            balance_stock_carton INTEGER,
            mc_to_give INTEGER,
            can_fulfill_carton INTEGER,
            shortfall INTEGER,
            status TEXT,
            required_kg REAL,
            remark TEXT,
            FOREIGN KEY (session_id) REFERENCES processing_sessions (id),
            FOREIGN KEY (order_file_id) REFERENCES uploaded_files (id)
        )
    ''')

    # Add remark column if it doesn't exist (for existing databases)
    try:
        cursor.execute("ALTER TABLE processing_results ADD COLUMN remark TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        # Column already exists
        pass

    # Create scheduled orders table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS scheduled_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            order_file_id INTEGER NOT NULL,
            scheduled_on TEXT NOT NULL,
            UNIQUE(session_id, order_file_id),
            FOREIGN KEY (session_id) REFERENCES processing_sessions (id),
            FOREIGN KEY (order_file_id) REFERENCES uploaded_files (id)
        )
    ''')

    # Create fish decisions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS fish_decisions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            fish_name TEXT NOT NULL,
            packed_size TEXT NOT NULL,
            order_name TEXT NOT NULL,
            decision TEXT NOT NULL,
                          UNIQUE(session_id, fish_name, packed_size, order_name),
              FOREIGN KEY (session_id) REFERENCES processing_sessions (id)
         )
     ''')
     
    # Create finished orders table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS finished_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id INTEGER NOT NULL,
            batch_token TEXT NOT NULL,
            order_token TEXT NOT NULL,
            order_name TEXT NOT NULL,
            weight REAL NOT NULL,
            finished_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(session_id, order_token),
            FOREIGN KEY (session_id) REFERENCES processing_sessions (id)
        )
    ''')
     
    # Create file comparison history table
    cursor.execute('''
         CREATE TABLE IF NOT EXISTS file_comparison_history (
             id INTEGER PRIMARY KEY AUTOINCREMENT,
             session_id INTEGER,
             order_file_id INTEGER,
             original_token TEXT,
             batch_token TEXT,
             comparison_data TEXT,
             changes_applied TEXT,
             comparison_summary TEXT,
             bangkok_datetime TEXT,
             created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
             FOREIGN KEY (session_id) REFERENCES processing_sessions(id),
             FOREIGN KEY (order_file_id) REFERENCES uploaded_files(id)
         )
    ''')
    
    # Check if order_file_id column exists, if not add it (migration)
    try:
        cursor.execute("SELECT order_file_id FROM processing_results LIMIT 1")
    except sqlite3.OperationalError:
        # Column doesn't exist, add it
        cursor.execute("ALTER TABLE processing_results ADD COLUMN order_file_id INTEGER")
        print("Added order_file_id column to processing_results table")
    
    # Check if new columns exist in uploaded_files table, if not add them (migration)
    try:
        cursor.execute("SELECT is_revision, revision_note FROM uploaded_files LIMIT 1")
    except sqlite3.OperationalError:
        # Columns don't exist, add them
        cursor.execute("ALTER TABLE uploaded_files ADD COLUMN is_revision BOOLEAN DEFAULT 0")
        cursor.execute("ALTER TABLE uploaded_files ADD COLUMN revision_note TEXT")
        print("Added revision tracking columns to uploaded_files table")
    
    conn.commit()
    conn.close()

def save_processing_session(session_token: str, batch_token: str = None, processing_type: str = 'single', summary: dict = None) -> int:
    """Save a processing session to database and return session ID."""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    # Bangkok time if available
    created_at = None
    try:
        if ZoneInfo:
            created_at = datetime.now(ZoneInfo('Asia/Bangkok')).strftime('%Y-%m-%d %H:%M:%S')
    except Exception:
        created_at = None

    if created_at:
        cursor.execute('''
            INSERT INTO processing_sessions 
            (session_token, batch_token, processing_type, total_items, full_items, not_full_items, not_have_items, 
             total_kg, full_kg, not_full_kg, not_have_kg, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            session_token,
            batch_token,
            processing_type,
            summary.get('total_items', 0) if summary else 0,
            summary.get('full', 0) if summary else 0,
            summary.get('not_full', 0) if summary else 0,
            summary.get('not_have', 0) if summary else 0,
            summary.get('total_kg_all', 0) if summary else 0,
            summary.get('total_kg_full', 0) if summary else 0,
            summary.get('total_kg_not_full', 0) if summary else 0,
            summary.get('total_kg_not_have', 0) if summary else 0,
            created_at,
        ))
    else:
        cursor.execute('''
            INSERT INTO processing_sessions 
            (session_token, batch_token, processing_type, total_items, full_items, not_full_items, not_have_items, 
             total_kg, full_kg, not_full_kg, not_have_kg)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            session_token,
            batch_token,
            processing_type,
            summary.get('total_items', 0) if summary else 0,
            summary.get('full', 0) if summary else 0,
            summary.get('not_full', 0) if summary else 0,
            summary.get('not_have', 0) if summary else 0,
            summary.get('total_kg_all', 0) if summary else 0,
            summary.get('total_kg_full', 0) if summary else 0,
            summary.get('total_kg_not_full', 0) if summary else 0,
            summary.get('total_kg_not_have', 0) if summary else 0
        ))
    
    session_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return session_id

def upsert_schedule(session_id: int, order_file_id: int, scheduled_on: str) -> None:
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO scheduled_orders (session_id, order_file_id, scheduled_on)
        VALUES (?, ?, ?)
        ON CONFLICT(session_id, order_file_id) DO UPDATE SET scheduled_on = excluded.scheduled_on
    ''', (session_id, order_file_id, scheduled_on))
    conn.commit()
    conn.close()

def save_file_comparison_history(session_id: int, order_file_id: int, original_token: str, 
                                batch_token: str, comparison_data: dict, changes_applied: list) -> None:
    """Save file comparison history to database with Bangkok timezone."""
    from datetime import datetime, timezone, timedelta
    
    # Get Bangkok timezone (+7 hours from UTC)
    bangkok_tz = timezone(timedelta(hours=7))
    bangkok_time = datetime.now(bangkok_tz)
    bangkok_datetime_str = bangkok_time.strftime('%Y-%m-%d %H:%M:%S %Z')
    
    # Prepare comparison summary
    summary = comparison_data.get('summary', {})
    comparison_summary = {
        'added': summary.get('added', 0),
        'modified': summary.get('modified', 0),
        'deleted': summary.get('deleted', 0),
        'unchanged': summary.get('unchanged', 0)
    }
    
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute(
        '''INSERT INTO file_comparison_history 
           (session_id, order_file_id, original_token, batch_token, comparison_data, 
            changes_applied, comparison_summary, bangkok_datetime)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
        (
            session_id, order_file_id, original_token, batch_token,
            json.dumps(comparison_data), json.dumps(changes_applied), json.dumps(comparison_summary),
            bangkok_datetime_str
        )
    )
    conn.commit()
    conn.close()

def get_file_comparison_history(batch_token: str) -> list:
    """Get file comparison history for a batch."""
    session_id = get_session_id_by_batch_token(batch_token)
    if not session_id:
        return []
    
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT fch.*, uf.original_filename
        FROM file_comparison_history fch
        JOIN uploaded_files uf ON fch.order_file_id = uf.id
        WHERE fch.session_id = ?
        ORDER BY fch.created_at DESC
    ''', (session_id,))
    
    history = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return history

def delete_schedule(session_id: int, order_file_id: int) -> None:
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM scheduled_orders WHERE session_id = ? AND order_file_id = ?', (session_id, order_file_id))
    conn.commit()
    conn.close()

def upsert_fish_decision(session_id: int, fish_name: str, packed_size: str, order_name: str, decision: str) -> None:
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO fish_decisions (session_id, fish_name, packed_size, order_name, decision)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(session_id, fish_name, packed_size, order_name) DO UPDATE SET decision = excluded.decision
    ''', (session_id, fish_name, packed_size, order_name, decision))
    conn.commit()
    conn.close()

def save_uploaded_file(session_id: int, file_type: str, original_filename: str, file_storage, is_revision: bool = False, revision_note: str = None) -> str:
    """Save uploaded file to disk and record in database."""
    # Generate unique filename
    file_ext = os.path.splitext(original_filename)[1]
    stored_filename = f"{uuid.uuid4().hex}{file_ext}"
    file_path = os.path.join(UPLOAD_FOLDER, stored_filename)
    
    # Save file to disk
    file_storage.save(file_path)
    file_size = os.path.getsize(file_path)
    
    # Record in database
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO uploaded_files (session_id, file_type, original_filename, stored_filename, file_size, is_revision, revision_note)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (session_id, file_type, original_filename, stored_filename, file_size, is_revision, revision_note))
    
    conn.commit()
    conn.close()
    
    return stored_filename

def save_revised_file(session_id: int, original_filename: str, file_path: str, revision_note: str = None) -> int:
    """Save a revised file that was created programmatically (not uploaded)."""
    import os
    
    # Generate unique stored filename
    file_ext = os.path.splitext(original_filename)[1]
    stored_filename = f"{uuid.uuid4().hex}{file_ext}"
    new_file_path = os.path.join(UPLOAD_FOLDER, stored_filename)
    
    # Copy the file to uploads folder with new name
    shutil.copy2(file_path, new_file_path)
    file_size = os.path.getsize(new_file_path)
    
    # Record in database as revision
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO uploaded_files (session_id, file_type, original_filename, stored_filename, file_size, is_revision, revision_note)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (session_id, 'order', original_filename, stored_filename, file_size, True, revision_note or 'Revised from Editor'))
    
    file_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return file_id

def save_processing_results(session_id: int, results: List[Dict[str, Any]], order_file_id: int = None):
    """Save processing results to database."""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    for result in results:
        cursor.execute('''
            INSERT INTO processing_results
            (session_id, order_file_id, fish_name, packed_size, order_carton, stock_carton, order_kg_per_ctn,
             stock_kg_per_ctn, balance_stock_carton, mc_to_give, can_fulfill_carton,
             shortfall, status, required_kg, remark)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            session_id,
            order_file_id,
            result.get('fish_name', ''),
            result.get('packed_size', ''),
            result.get('order_carton', 0),
            result.get('stock_carton', 0),
            result.get('order_kg_per_ctn', 0),
            result.get('stock_kg_per_ctn', 0),
            result.get('balance_stock_carton', 0),
            result.get('mc_to_give', 0),
            result.get('can_fulfill_carton', 0),
            result.get('shortfall', 0),
            result.get('status', ''),
            result.get('required_kg', 0),
            result.get('remark', '')
        ))
    
    conn.commit()
    conn.close()

def get_recent_sessions(limit: int = 10) -> List[Dict[str, Any]]:
    """Get recent processing sessions for display."""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT ps.*, 
               COUNT(CASE WHEN uf.file_type = 'stock' THEN 1 END) as stock_files,
               COUNT(CASE WHEN uf.file_type = 'order' THEN 1 END) as order_files,
               GROUP_CONCAT(CASE WHEN uf.file_type = 'stock' THEN uf.original_filename END) as stock_filename,
               GROUP_CONCAT(CASE WHEN uf.file_type = 'order' THEN uf.original_filename END, ', ') as order_filenames
        FROM processing_sessions ps
        LEFT JOIN uploaded_files uf ON ps.id = uf.session_id
        GROUP BY ps.id
        ORDER BY datetime(ps.created_at) DESC
        LIMIT ?
    ''', (limit,))
    
    sessions = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return sessions

def get_finished_orders_for_batch(batch_token: str) -> List[Dict[str, Any]]:
    """Return list of finished orders for a batch token."""
    finished_orders = []
    session_id = get_session_id_by_batch_token(batch_token)
    if not session_id:
        return finished_orders

    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute('''
        SELECT order_token, order_name, weight, finished_at
        FROM finished_orders
        WHERE session_id = ?
        ORDER BY finished_at DESC
    ''', (session_id,))

    for row in cursor.fetchall():
        finished_orders.append({
            'token': row['order_token'],
            'name': row['order_name'],
            'weight': row['weight'],
            'finished_at': row['finished_at']
        })

    conn.close()

def get_session_files(session_id: int) -> Dict[str, List[Dict[str, Any]]]:
    """Get detailed file information for a specific session."""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute('''
        SELECT file_type, original_filename, stored_filename, file_size,
               is_revision, revision_note, uploaded_at
        FROM uploaded_files
        WHERE session_id = ?
        ORDER BY uploaded_at ASC
    ''', (session_id,))

    files = {'stock': [], 'import_stock': [], 'extra_load_stock': [], 'order': []}
    for row in cursor.fetchall():
        file_info = {
            'original_filename': row['original_filename'],
            'stored_filename': row['stored_filename'],
            'file_size': row['file_size'],
            'is_revision': bool(row['is_revision']),
            'revision_note': row['revision_note'],
            'uploaded_at': row['uploaded_at']
        }
        files[row['file_type']].append(file_info)

    conn.close()
    return files

def get_session_id_by_batch_token(batch_token: str) -> Optional[int]:
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM processing_sessions WHERE batch_token = ? LIMIT 1", (batch_token,))
    row = cursor.fetchone()
    conn.close()
    return int(row[0]) if row else None

def get_schedules_for_batch(batch_token: str) -> Dict[str, str]:
    """Return mapping of order original_filename -> scheduled_on for a batch token."""
    mapping: Dict[str, str] = {}
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('''
        SELECT uf.original_filename AS order_name, so.scheduled_on
        FROM processing_sessions ps
        JOIN uploaded_files uf ON uf.session_id = ps.id AND uf.file_type = 'order'
        LEFT JOIN scheduled_orders so ON so.session_id = ps.id AND so.order_file_id = uf.id
        WHERE ps.batch_token = ?
    ''', (batch_token,))
    for row in cursor.fetchall():
        if row['scheduled_on']:
            mapping[str(row['order_name'])] = str(row['scheduled_on'])
    conn.close()
    return mapping

def get_session_by_id(session_id: int) -> Optional[Dict[str, Any]]:
    """Get a specific session with its files and results."""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get session info
    cursor.execute('''
        SELECT ps.*, 
               COUNT(CASE WHEN uf.file_type = 'stock' THEN 1 END) as stock_files,
               COUNT(CASE WHEN uf.file_type = 'order' THEN 1 END) as order_files,
               GROUP_CONCAT(CASE WHEN uf.file_type = 'stock' THEN uf.original_filename END) as stock_filename,
               GROUP_CONCAT(CASE WHEN uf.file_type = 'order' THEN uf.original_filename END, ', ') as order_filenames
        FROM processing_sessions ps
        LEFT JOIN uploaded_files uf ON ps.id = uf.session_id
        WHERE ps.id = ?
        GROUP BY ps.id
    ''', (session_id,))
    
    session = cursor.fetchone()
    if not session:
        conn.close()
        return None
    
    session = dict(session)
    
    # Get processing results grouped by order file for batch sessions
    if session.get('processing_type') == 'batch':
        # Get all uploaded files for this session
        cursor.execute('''
            SELECT * FROM uploaded_files
            WHERE session_id = ? AND file_type = 'order'
            ORDER BY id
        ''', (session_id,))
        
        order_files = [dict(row) for row in cursor.fetchall()]
        session['order_files_list'] = order_files
        
        # Get results grouped by order file (we'll need to reconstruct this)
        cursor.execute('''
            SELECT * FROM processing_results
            WHERE session_id = ?
            ORDER BY id
        ''', (session_id,))
        
        all_results = [dict(row) for row in cursor.fetchall()]
        session['all_results'] = all_results
    else:
        # Get processing results for single sessions
        cursor.execute('''
            SELECT * FROM processing_results
            WHERE session_id = ?
            ORDER BY id
        ''', (session_id,))
        
        results = [dict(row) for row in cursor.fetchall()]
        session['results'] = results
    
    conn.close()
    return session

def restore_batch_session_to_memory(session: Dict[str, Any]) -> str:
    """Restore a batch session to memory stores and return batch_token."""
    if not session.get('batch_token'):
        return None
    
    batch_token = session['batch_token']
    
    # Check if batch is already in memory with updated data - if so, just return the token
    if batch_token in BATCH_STORE and BATCH_STORE[batch_token]:
        # Verify all tokens still exist in RESULT_STORE
        tokens_valid = all(token in RESULT_STORE for token in BATCH_STORE[batch_token])
        if tokens_valid:
            return batch_token  # Return existing batch token without reloading
    
    stock_name = session.get('stock_filename', 'Unknown Stock')
    
    # Get order files and reprocess them from the actual stored files
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get stock file first
    cursor.execute('''
        SELECT stored_filename FROM uploaded_files
        WHERE session_id = ? AND file_type = 'stock'
        LIMIT 1
    ''', (session['id'],))
    stock_file_row = cursor.fetchone()
    
    if not stock_file_row:
        conn.close()
        return None
    
    # Load stock data from saved file
    stock_file_path = os.path.join(UPLOAD_FOLDER, stock_file_row['stored_filename'])
    try:
        with open(stock_file_path, 'rb') as f:
            file_like = io.BytesIO(f.read())
            file_like.filename = stock_file_path
            stock_ds = load_tabular(file_like)
    except Exception as e:
        print(f"Error loading stock file: {e}")
        conn.close()
        return None
    
    # Get order files
    cursor.execute('''
        SELECT id, original_filename, stored_filename FROM uploaded_files
        WHERE session_id = ? AND file_type = 'order'
        ORDER BY id
    ''', (session['id'],))
    order_files = cursor.fetchall()
    
    # Get processed results from database instead of reprocessing files
    files_data = {}
    for order_file in order_files:
        try:
            # Get processed results from database for this order file
            cursor.execute('''
                SELECT * FROM processing_results
                WHERE session_id = ? AND order_file_id = ?
                ORDER BY id DESC
            ''', (session['id'], order_file['id']))
            
            db_results = [dict(row) for row in cursor.fetchall()]
            
            # Convert database results to the expected format
            result_rows = []
            for db_result in db_results:
                result_row = {
                    'fish_name': db_result.get('fish_name', ''),
                    'packed_size': db_result.get('packed_size', ''),
                    'order_carton': db_result.get('order_carton', 0),
                    'stock_carton': db_result.get('stock_carton', 0),
                    'order_kg_per_ctn': db_result.get('order_kg_per_ctn', 0),
                    'stock_kg_per_ctn': db_result.get('stock_kg_per_ctn', 0),
                    'balance_stock_carton': db_result.get('balance_stock_carton', 0),
                    'mc_to_give': db_result.get('mc_to_give', 0),
                    'can_fulfill_carton': db_result.get('can_fulfill_carton', 0),
                    'shortfall': db_result.get('shortfall', 0),
                    'status': db_result.get('status', ''),
                    'required_kg': db_result.get('required_kg', 0),
                    'matched_by': 'Historical Data'  # Add indicator this is from DB
                }
                result_rows.append(result_row)
            
            files_data[order_file['id']] = {
                'filename': order_file['original_filename'],
                'results': result_rows,
                'order_data': []  # We don't need original order data for historical sessions
            }
            
        except Exception as e:
            print(f"Warning: Failed to load database results for order file {order_file['original_filename']}: {e}")
            # Fallback to empty results
            files_data[order_file['id']] = {
                'filename': order_file['original_filename'],
                'results': [],
                'order_data': []
            }
    
    conn.close()
    
    # Create tokens for each order file and populate RESULT_STORE
    token_list = []
    
    for file_id, file_data in files_data.items():
        token = uuid.uuid4().hex
        order_name = file_data['filename']
        result_rows = file_data['results']
        
        # Calculate summary for this order file
        def sum_required_kg(rows):
            return round(sum(float(r.get('required_kg', 0) or 0) for r in rows), 3)
        
        def sum_ready_kg(rows):
            return round(sum(float(r.get('stock_carton', 0) or 0) * float(r.get('stock_kg_per_ctn', 0) or 0) for r in rows), 3)
        
        def sum_fulfillable_kg(rows):
            return round(sum(float(r.get('can_fulfill_carton', 0) or 0) * float(r.get('order_kg_per_ctn', 0) or 0) for r in rows), 3)
        
        summary = {
            'total_items': len(result_rows),
            'full': sum(1 for r in result_rows if r.get('status') == 'Full'),
            'not_full': sum(1 for r in result_rows if r.get('status') == 'Not Full'),
            'not_have': sum(1 for r in result_rows if r.get('status') == 'Not have'),
            'total_kg_all': sum_required_kg(result_rows),
            'total_kg_full': sum_required_kg([r for r in result_rows if r.get('status') == 'Full']),
            'total_kg_not_full': sum_required_kg([r for r in result_rows if r.get('status') == 'Not Full']),
            'total_kg_not_have': sum_required_kg([r for r in result_rows if r.get('status') == 'Not have']),
            'ready_kg': sum_ready_kg(result_rows),
            'fulfillable_kg': sum_fulfillable_kg(result_rows)
        }
        
        # Create Excel/PDF bytes for historical data
        excel_bytes = rows_to_excel_bytes(result_rows)
        pdf_bytes = rows_to_pdf_bytes(result_rows)
        
        # Store in RESULT_STORE
        order_basename = os.path.splitext(order_name)[0]
        
        # Use the actual order data from the reloaded file
        order_data = file_data['order_data']
        original_order_data = []
        for row in order_data:
            normalized = try_map_row(row)
            weight_mc = normalized.get('weight_mc', '')
            
            # First try to parse order_kg_per_ctn from weight_mc
            try:
                order_kg_per_ctn = float(weight_mc) if weight_mc not in (None, "") else 0.0
            except Exception:
                order_kg_per_ctn = parse_kg_per_carton(weight_mc)
            
            # If we still don't have weight_mc but we parsed order_kg_per_ctn from packed size, use it
            if not weight_mc and order_kg_per_ctn == 0:
                try:
                    order_kg_per_ctn = parse_kg_per_carton(normalized.get('packed size', ''))
                    if order_kg_per_ctn > 0:
                        weight_mc = str(order_kg_per_ctn)
                except:
                    pass
            elif not weight_mc and order_kg_per_ctn > 0:
                weight_mc = str(order_kg_per_ctn)
            
            order_row = {
                'fish name': normalized.get('fish name', ''),
                'packed size': normalized.get('packed size', ''),
                'pack': normalized.get('pack', ''),
                'total carton': to_int(normalized.get('total carton', 0)),
                'weight_mc': weight_mc,
                'order_kg_per_ctn': order_kg_per_ctn,
                'remark': normalized.get('remark', '')
            }
            original_order_data.append(order_row)
        
        RESULT_STORE[token] = {
            "excel": excel_bytes,
            "pdf": pdf_bytes,
            "excel_name": f"{order_basename} Calculation.xlsx",
            "pdf_name": f"{order_basename} Calculation.pdf",
            "rows_json": io.BytesIO(str(result_rows).encode("utf-8")).getvalue(),
            "summary_json": io.BytesIO(str(summary).encode("utf-8")).getvalue(),
            "stock_name": stock_name.encode("utf-8"),
            "order_name": order_name.encode("utf-8"),
            "original_order_json": io.BytesIO(str(original_order_data).encode("utf-8")).getvalue(),
        }
        
        token_list.append(token)
    
    # Store in BATCH_STORE
    BATCH_STORE[batch_token] = token_list
    
    return batch_token

def get_session_by_token(session_token: str) -> Optional[Dict[str, Any]]:
    """Get a session by its token."""
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute('SELECT id FROM processing_sessions WHERE session_token = ?', (session_token,))
    result = cursor.fetchone()
    conn.close()
    
    if result:
        return get_session_by_id(result['id'])
    return None

# Initialize database on startup
init_database()

# ----- Data Models -----
@dataclass
class Dataset:
    name: str
    rows: List[Dict[str, Any]]
    sheet_names: List[str]


# ----- Helpers -----
NORMALIZE_MAP = {
    "fish name": ["fish name", "fish", "product", "product name", "name"],
    "packed size": ["packed size", "pack", "pack size", "size"],
    "pack": ["pack", "packed size", "pack size", "size"],
    "total carton": [
        "total carton",
        "total_ctn",
        "total ctn",
        "ctn",
        "carton",
        "cartons",
        "qty",
        "quantity",
    ],
    "weight_mc": [
        "weight_mc",
        "weight mt",
        "weight_mt",
        "net_weigh",
        "net weight",
        "weight per mc",
        "mc_weight",
        "kg/ctn",
        "kg per ctn",
        "weight mc",
        "order kg/ctn",
    ],
    "remark": ["remark", "remarks", "note", "notes", "comment", "comments"],
}


def load_excel(file_storage, preferred_sheet: Optional[str] = None) -> Dataset:
    # Reset pointer and load workbook
    try:
        file_storage.stream.seek(0)
    except Exception:  # noqa: BLE001
        pass
    wb = load_workbook(file_storage, data_only=True)
    sheet_name = preferred_sheet or wb.sheetnames[0]
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        headers = []
    headers = [str(h).strip() if h is not None else "" for h in headers]

    rows: List[Dict[str, Any]] = []
    for row in rows_iter:
        record = {}
        empty = True
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            if val not in (None, ""):
                empty = False
            record[h] = val
        if not empty:
            rows.append(record)

    return Dataset(name=getattr(file_storage, "filename", "uploaded.xlsx"), rows=rows, sheet_names=wb.sheetnames)


def load_csv(file_storage) -> Dataset:
    try:
        file_storage.stream.seek(0)
    except Exception:  # noqa: BLE001
        pass
    raw = file_storage.read()
    try:
        text = raw.decode("utf-8-sig")
    except Exception:  # noqa: BLE001
        text = raw.decode("latin-1", errors="ignore")
    reader = csv.DictReader(text.splitlines())
    rows: List[Dict[str, Any]] = []
    for r in reader:
        # Drop None keys if any
        rows.append({(k or ""): v for k, v in r.items()})
    return Dataset(name=getattr(file_storage, "filename", "uploaded.csv"), rows=rows, sheet_names=["CSV"])


def load_tabular(file_storage) -> Dataset:
    filename = (getattr(file_storage, "filename", "") or "").lower()
    if filename.endswith(".csv"):
        return load_csv(file_storage)
    # default to excel
    return load_excel(file_storage)


def try_map_row(row: Dict[str, Any]) -> Dict[str, Any]:
    lower_map = {k.lower().strip(): k for k in row.keys()}
    normalized: Dict[str, Any] = {}
    for key, aliases in NORMALIZE_MAP.items():
        value = None
        for alias in aliases:
            if alias in lower_map:
                value = row.get(lower_map[alias])
                break
        normalized[key] = value
    return normalized


def normalize_text_val(val: Any) -> str:
    if val is None:
        return ""
    return str(val).strip().upper().replace("  ", " ")


_PARENS_RE = re.compile(r"\([^\)]*\)")
_MULTISPACE_RE = re.compile(r"\s+")
_RANGE_RE = re.compile(r"(\d+)\s*[-–]\s*(\d+)")
_PUNCT_BREAK_RE = re.compile(r"[\./_,]+")
_NON_ALNUM_RE = re.compile(r"[^A-Z0-9]+")


def canonicalize_product(text: Any) -> str:
    s = normalize_text_val(text)
    s = _PARENS_RE.sub(" ", s)
    s = _RANGE_RE.sub(lambda m: f"{m.group(1)} {m.group(2)}", s)
    # Ignore all special characters
    s = _NON_ALNUM_RE.sub(" ", s)
    # remove unit/noise tokens
    stop = {
        "G", "GM", "GMS", "GRAM", "GRAMS", "PCS", "PC", "CTN", "CTNS", "GL", "GLAZE",
        "WITH", "PRINT", "BAG", "RIDER", "STICKER", "PACK", "SIZE", "KG",
        # common glaze percentages as numbers
        "5", "10", "15", "20", "25", "30", "35", "40",
    }
    tokens = [t for t in _MULTISPACE_RE.split(s) if t]
    filtered: List[str] = []
    for t in tokens:
        if t in stop:
            continue
        filtered.append(t)
    # Return a compact key without spaces so 'SILVER CARP' == 'SILVERCARP'
    return "".join(filtered)


def canonicalize_pack(text: Any) -> str:
    s = normalize_text_val(text)
    s = _RANGE_RE.sub(lambda m: f"{m.group(1)} {m.group(2)}", s)
    # Ignore all special characters
    s = _NON_ALNUM_RE.sub(" ", s)
    s = _MULTISPACE_RE.sub(" ", s)
    # Return compact form (no spaces) to ignore spacing differences
    return s.replace(" ", "").strip()


_MASS_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(KG|KGS|G|GM|GRAM|GRAMS)")
_MULT_X_RE = re.compile(r"X\s*(\d+)")


def parse_kg_per_carton(text: Any) -> float:
    """Extract approximate kg per carton from a pack description, e.g. '1 KG X 10 BAG/CTN' -> 10.0.
    Returns 0.0 if cannot parse.
    """
    s = normalize_text_val(text)
    if not s:
        return 0.0
    s = s.replace("×", "X")
    # Combine first mass block with a nearby multiplier if present
    match = _MASS_RE.search(s)
    if not match:
        return 0.0
    value = float(match.group(1))
    unit = match.group(2)
    kg = value if unit.startswith("K") else value / 1000.0
    # Look ahead for a multiplier within the next ~20 chars
    tail = s[match.end():match.end() + 30]
    m2 = _MULT_X_RE.search(tail)
    mult = float(m2.group(1)) if m2 else 1.0
    return kg * mult


def to_int(val: Any) -> int:
    try:
        if val is None or val == "":
            return 0
        return int(float(val))
    except Exception:  # noqa: BLE001
        return 0


def compute_matches(stock_rows: List[Dict[str, Any]], order_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    # Build stock lookups
    stock_by_prod_pack: Dict[Tuple[str, str], Dict[str, Any]] = {}
    stock_by_prod: Dict[str, Dict[str, Any]] = {}

    for r in stock_rows:
        nr = try_map_row(r)
        prod_key = canonicalize_product(nr.get("fish name"))
        pack_key = canonicalize_pack(nr.get("packed size"))
        qty = to_int(nr.get("total carton"))
        kg_per_ctn = parse_kg_per_carton(nr.get("packed size"))
        if not prod_key and not pack_key:
            continue
        if prod_key:
            agg = stock_by_prod.get(prod_key, {"ctn": 0, "kg_per_ctn": kg_per_ctn})
            agg["ctn"] += qty
            # prefer non-zero kg_per_ctn when available
            if agg.get("kg_per_ctn", 0) == 0 and kg_per_ctn:
                agg["kg_per_ctn"] = kg_per_ctn
            stock_by_prod[prod_key] = agg
        key = (prod_key, pack_key)
        agg2 = stock_by_prod_pack.get(key, {"ctn": 0, "kg_per_ctn": kg_per_ctn})
        agg2["ctn"] += qty
        if agg2.get("kg_per_ctn", 0) == 0 and kg_per_ctn:
            agg2["kg_per_ctn"] = kg_per_ctn
        stock_by_prod_pack[key] = agg2

    results: List[Dict[str, Any]] = []
    for r in order_rows:
        nr = try_map_row(r)
        fish_text = nr.get("fish name")
        pack_text = nr.get("packed size")
        prod_key = canonicalize_product(fish_text)
        pack_key = canonicalize_pack(pack_text)
        order_qty = to_int(nr.get("total carton"))
        weight_mc = nr.get("weight_mc")
        try:
            order_kg_per_ctn = float(weight_mc) if weight_mc not in (None, "") else 0.0
        except Exception:  # noqa: BLE001
            order_kg_per_ctn = parse_kg_per_carton(weight_mc)

        matched_by = ""
        stock_qty = 0
        # Prefer exact product+pack match when pack is present
        if pack_key and (prod_key, pack_key) in stock_by_prod_pack:
            stock_info = stock_by_prod_pack.get((prod_key, pack_key), {"ctn": 0, "kg_per_ctn": 0.0})
            stock_qty = int(stock_info.get("ctn", 0))
            stock_pack_kg = float(stock_info.get("kg_per_ctn", 0.0))
            matched_by = "product+pack"
        elif prod_key in stock_by_prod:
            stock_info = stock_by_prod.get(prod_key, {"ctn": 0, "kg_per_ctn": 0.0})
            stock_qty = int(stock_info.get("ctn", 0))
            stock_pack_kg = float(stock_info.get("kg_per_ctn", 0.0))
            matched_by = "product"
        else:
            stock_pack_kg = 0.0

        # Weight-aware calculations
        effective_stock_ctn = stock_qty
        mc_to_give = 0
        if order_kg_per_ctn and stock_pack_kg:
            total_stock_kg = stock_qty * stock_pack_kg
            # how many order-equivalent cartons can stock satisfy
            effective_stock_ctn = int(total_stock_kg // max(order_kg_per_ctn, 0.0001))
            # MC to pick from stock to satisfy order requirement
            required_kg = order_qty * order_kg_per_ctn
            mc_needed = math.ceil(required_kg / max(stock_pack_kg, 0.0001))
            mc_to_give = min(stock_qty, mc_needed)
        else:
            mc_to_give = min(stock_qty, order_qty)

        if effective_stock_ctn <= 0:
            status = "Not have"
        elif effective_stock_ctn < order_qty:
            status = "Not Full"
        else:
            status = "Full"

        fulfilled_ctn = min(order_qty, effective_stock_ctn)
        # Balance stock after giving the computed MC from stock
        balance_after_order = max(stock_qty - mc_to_give, 0)

        result = {
            "fish_name": fish_text,
            "packed_size": pack_text,
            "order_carton": order_qty,
            "stock_carton": stock_qty,  # raw MC from stock file
            "can_fulfill_carton": fulfilled_ctn,
            "shortfall": max(order_qty - effective_stock_ctn, 0),
            "status": status,
            "matched_by": matched_by,
            "order_kg_per_ctn": round(order_kg_per_ctn, 3) if order_kg_per_ctn else 0,
            "stock_kg_per_ctn": round(stock_pack_kg, 3) if stock_pack_kg else 0,
            "balance_stock_carton": balance_after_order,
            "mc_to_give": mc_to_give,
            "required_kg": round(order_qty * (order_kg_per_ctn or 0), 3),
            "remark": nr.get('remark', '')
        }
        results.append(result)

    return results


def rows_to_excel_bytes(rows: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()

    headers = [
        "Fish Name",
        "Packed Size",
        "Order CTN",
        "Stock CTN",
        "Can Fulfill",
        "Shortfall",
        "Status",
        "Order KG/CTN",
        "Stock KG/CTN",
        "Balance Stock CTN",
    ]

    def write_sheet(ws, data_rows: List[Dict[str, Any]]):
        ws.append(headers)
        for r in data_rows:
            ws.append([
                r.get("fish_name", ""),
                r.get("packed_size", ""),
                r.get("order_carton", 0),
                r.get("stock_carton", 0),
                r.get("can_fulfill_carton", 0),
                r.get("shortfall", 0),
                r.get("status", ""),
                r.get("order_kg_per_ctn", 0),
                r.get("stock_kg_per_ctn", 0),
                r.get("balance_stock_carton", 0),
            ])
        ws.freeze_panes = "A2"
        widths = [35, 18, 12, 12, 14, 12, 12, 14, 14, 18]
        for idx, width in enumerate(widths, start=1):
            col = ws.cell(row=1, column=idx).column_letter
            ws.column_dimensions[col].width = width

    # Create sheets: All, Full, Not Full, Not have
    ws_all = wb.active
    ws_all.title = "All"
    write_sheet(ws_all, rows)

    ws_full = wb.create_sheet(title="Full")
    write_sheet(ws_full, [r for r in rows if r.get("status") == "Full"])

    ws_nf = wb.create_sheet(title="Not Full")
    write_sheet(ws_nf, [r for r in rows if r.get("status") == "Not Full"])

    ws_nh = wb.create_sheet(title="Not have")
    write_sheet(ws_nh, [r for r in rows if r.get("status") == "Not have"])

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def rows_to_pdf_bytes(rows: List[Dict[str, Any]]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
    from reportlab.platypus.tableofcontents import TableOfContents
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus.frames import Frame
    from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
    from reportlab.platypus.flowables import PageBreak
    import os

    buffer = io.BytesIO()
    
    # Custom function to draw header and footer
    def draw_page_decorations(canvas, doc):
        pagesize = landscape(A3)
        
        # Header with images
        canvas.saveState()
        
        # Draw header images if they exist
        oac_path = os.path.join(os.path.dirname(__file__), 'static', 'images', 'logo_OAC.png')
        logo_thai_path = os.path.join(os.path.dirname(__file__), 'static', 'images', 'logo-thai-bg.png')
        
        # Left header image (OAC.png)
        if os.path.exists(oac_path):
            try:
                canvas.drawImage(oac_path, 30*mm, pagesize[1]-30*mm, width=40*mm, height=20*mm, preserveAspectRatio=True)
            except Exception:
                pass
        
        # Right header image (logo-thai.png)  
        if os.path.exists(logo_thai_path):
            try:
                canvas.drawImage(logo_thai_path, pagesize[0]-70*mm, pagesize[1]-30*mm, width=40*mm, height=20*mm, preserveAspectRatio=True)
            except Exception:
                pass
        
        # Header title in center
        canvas.setFont("Helvetica-Bold", 16)
        canvas.drawCentredString(pagesize[0]/2, pagesize[1]-20*mm, "C.K. Frozen Fish & Food Co.,Ltd (Chachoengsao)")
        canvas.drawCentredString(pagesize[0]/2, pagesize[1]-30*mm, "Order Availability Result")
        # Header line
        canvas.setStrokeColor(colors.grey)
        canvas.line(20*mm, pagesize[1]-35*mm, pagesize[0]-20*mm, pagesize[1]-35*mm)
        
        # Footer
        canvas.setFont("Helvetica", 9)
        
        # Company name on left
        canvas.drawString(25*mm, 10*mm, "C.K Thailand")
        
        # Page number on right
        page_num = canvas.getPageNumber()
        canvas.drawRightString(pagesize[0]-25*mm, 10*mm, f"Page {page_num}")
        
        # Footer line
        canvas.setStrokeColor(colors.grey)
        canvas.line(20*mm, 15*mm, pagesize[0]-20*mm, 15*mm)
        
        canvas.restoreState()

    # Use A3 landscape for more space with simple document
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A3), leftMargin=20*mm, rightMargin=20*mm, 
                          topMargin=40*mm, bottomMargin=20*mm)
    
    story = []
    story.append(Spacer(1, 10*mm))

    # Complete headers with all columns
    headers = [
        "Fish Name",
        "Packed Size",
        "Order CTN",
        "Stock CTN",
        "Order KG/CTN",
        "Stock KG/CTN",
        "Balance Stock CTN",
        "Can Fulfill",
        "Shortfall",
        "Required KG",
        "Status"
    ]
    
    # Build data rows
    styles = getSampleStyleSheet()
    
    # Define custom paragraph styles for text wrapping
    text_style = ParagraphStyle(
        'TableText',
        parent=styles['Normal'],
        fontSize=10,
        leading=12,
        alignment=TA_LEFT,
        wordWrap='CJK'
    )
    
    number_style = ParagraphStyle(
        'TableNumber',
        parent=styles['Normal'],
        fontSize=10,
        leading=12,
        alignment=TA_RIGHT,
        wordWrap='CJK'
    )
    
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=11,
        leading=14,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#2c3e50"),  # Dark blue-gray text
        fontName='Helvetica-Bold',
        wordWrap='CJK'
    )
    
    # Status column styles with colors
    status_full_style = ParagraphStyle(
        'StatusFull',
        parent=styles['Normal'],
        fontSize=10,
        leading=12,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#059669"),  # Green
        wordWrap='CJK'
    )
    
    status_partial_style = ParagraphStyle(
        'StatusPartial',
        parent=styles['Normal'],
        fontSize=10,
        leading=12,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#d97706"),  # Orange
        wordWrap='CJK'
    )
    
    status_none_style = ParagraphStyle(
        'StatusNone',
        parent=styles['Normal'],
        fontSize=10,
        leading=12,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#dc2626"),  # Red
        wordWrap='CJK'
    )
    
    # Convert headers to wrapped paragraphs
    wrapped_headers = [Paragraph(str(header), header_style) for header in headers]
    data = [wrapped_headers]
    
    # Initialize totals
    totals = {
        'order_carton': 0,
        'stock_carton': 0,
        'order_kg_per_ctn': 0,
        'stock_kg_per_ctn': 0,
        'balance_stock_carton': 0,
        'can_fulfill_carton': 0,
        'shortfall': 0,
        'required_kg': 0
    }
    
    for r in rows:
        # Add numeric values to totals
        totals['order_carton'] += float(r.get("order_carton", 0) or 0)
        totals['stock_carton'] += float(r.get("stock_carton", 0) or 0)
        totals['order_kg_per_ctn'] += float(r.get("order_kg_per_ctn", 0) or 0)
        totals['stock_kg_per_ctn'] += float(r.get("stock_kg_per_ctn", 0) or 0)
        totals['balance_stock_carton'] += float(r.get("balance_stock_carton", 0) or 0)
        totals['can_fulfill_carton'] += float(r.get("can_fulfill_carton", 0) or 0)
        totals['shortfall'] += float(r.get("shortfall", 0) or 0)
        totals['required_kg'] += float(r.get("required_kg", 0) or 0)
        
        # Determine status style based on status value
        status_text = str(r.get("status", "")).lower()
        if 'full' in status_text and 'not full' not in status_text:
            status_style = status_full_style
        elif 'not full' in status_text:
            status_style = status_partial_style
        elif 'not have' in status_text:
            status_style = status_none_style
        else:
            status_style = text_style
        
        # Create wrapped paragraphs for each cell
        row_data = [
            Paragraph(str(r.get("fish_name", "")), text_style),
            Paragraph(str(r.get("packed_size", "")), text_style),
            Paragraph(str(r.get("order_carton", 0)), number_style),
            Paragraph(str(r.get("stock_carton", 0)), number_style),
            Paragraph(str(round(float(r.get("order_kg_per_ctn", 0) or 0), 2)), number_style),
            Paragraph(str(round(float(r.get("stock_kg_per_ctn", 0) or 0), 2)), number_style),
            Paragraph(str(r.get("balance_stock_carton", 0)), number_style),
            Paragraph(str(r.get("can_fulfill_carton", 0)), number_style),
            Paragraph(str(r.get("shortfall", 0)), number_style),
            Paragraph(str(round(float(r.get("required_kg", 0) or 0), 2)), number_style),
            Paragraph(str(r.get("status", "")), status_style)
        ]
        data.append(row_data)

    # Add totals row with paragraphs
    totals_style = ParagraphStyle(
        'TotalsText',
        parent=styles['Normal'],
        fontSize=11,
        leading=14,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor("#495057"),  # Professional dark gray
        wordWrap='CJK'
    )
    
    totals_number_style = ParagraphStyle(
        'TotalsNumber',
        parent=styles['Normal'],
        fontSize=11,
        leading=14,
        alignment=TA_RIGHT,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor("#495057"),  # Professional dark gray
        wordWrap='CJK'
    )
    
    totals_row = [
        Paragraph("TOTAL", totals_style),
        Paragraph("", totals_style),
        Paragraph(str(round(totals['order_carton'])), totals_number_style),
        Paragraph(str(round(totals['stock_carton'])), totals_number_style),
        Paragraph(str(round(totals['order_kg_per_ctn'], 2)), totals_number_style),
        Paragraph(str(round(totals['stock_kg_per_ctn'], 2)), totals_number_style),
        Paragraph(str(round(totals['balance_stock_carton'])), totals_number_style),
        Paragraph(str(round(totals['can_fulfill_carton'])), totals_number_style),
        Paragraph(str(round(totals['shortfall'])), totals_number_style),
        Paragraph(str(round(totals['required_kg'], 2)), totals_number_style),
        Paragraph("", totals_style)
    ]
    data.append(totals_row)

    # Calculate column widths to fit A3 landscape
    page_width = landscape(A3)[0] - 40*mm  # Total width minus margins
    col_widths = [
        page_width * 0.20,  # Fish Name - 20%
        page_width * 0.15,  # Packed Size - 15%
        page_width * 0.08,  # Order CTN - 8%
        page_width * 0.08,  # Stock CTN - 8%
        page_width * 0.08,  # Order KG/CTN - 8%
        page_width * 0.08,  # Stock KG/CTN - 8%
        page_width * 0.09,  # Balance Stock CTN - 9%
        page_width * 0.08,  # Can Fulfill - 8%
        page_width * 0.07,  # Shortfall - 7%
        page_width * 0.08,  # Required KG - 8%
        page_width * 0.08   # Status - 8%
    ]

    # Set row heights for better visibility with text wrapping
    row_heights = [28] * len(data)  # 28 points = about 9.88mm height for all rows (much taller)
    row_heights[0] = 32  # Header row even taller (32 points = about 11.29mm)
    row_heights[-1] = 30  # Totals row (30 points = about 10.58mm)
    
    table = Table(data, colWidths=col_widths, rowHeights=row_heights, repeatRows=1)
    
    # Clean professional table style - matching your uploaded image colors
    table_style = [
        # Header row styling - clean light gray background
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8f9fa")),
        
        # Totals row styling - subtle blue-gray background
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#e9ecef")),
        
        # Grid and borders - clean gray lines
        ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#dee2e6")),
        ("LINEBELOW", (0, 0), (-1, 0), 2, colors.HexColor("#6c757d")),  # Header bottom border
        ("LINEABOVE", (0, -1), (-1, -1), 2, colors.HexColor("#6c757d")),  # Totals top border
        
        # Row backgrounds - clean alternating white and very light gray
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f8f9fa")]),
        
        # Vertical alignment
        ("VALIGN", (0, 0), (-1, -1), "TOP"),  # Top align for better text wrapping
        
        # Cell padding for better spacing
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    
    # Status colors are now handled by paragraph styles, no additional formatting needed
    
    table.setStyle(TableStyle(table_style))
    story.append(table)

    # Build document with custom header/footer function
    doc.build(story, onFirstPage=draw_page_decorations, onLaterPages=draw_page_decorations)
    return buffer.getvalue()


# ----- Routes -----
@app.get("/")
def index():
    recent_sessions = get_recent_sessions(limit=5)
    return render_template("index.html", recent_sessions=recent_sessions)


@app.post('/delete-session')
def delete_session():
    session_id = request.form.get('session_id', type=int)
    if not session_id:
        flash('Invalid session selected.', 'error')
        return redirect(url_for('index'))
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()
        # If batch, capture batch_token to also clear in-memory batch if present
        cursor.execute('SELECT batch_token FROM processing_sessions WHERE id=?', (session_id,))
        row = cursor.fetchone()
        batch_token = row[0] if row else None

        # Get all uploaded files for this session to delete from uploads folder
        cursor.execute('SELECT stored_filename FROM uploaded_files WHERE session_id=?', (session_id,))
        file_rows = cursor.fetchall()
        
        # Delete files from uploads folder
        for file_row in file_rows:
            stored_filename = file_row[0]
            if stored_filename:
                file_path = os.path.join(UPLOAD_FOLDER, stored_filename)
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                except Exception as e:
                    print(f"Warning: Failed to delete file {file_path}: {e}")

        # Delete related rows
        cursor.execute('DELETE FROM scheduled_orders WHERE session_id=?', (session_id,))
        cursor.execute('DELETE FROM fish_decisions WHERE session_id=?', (session_id,))
        cursor.execute('DELETE FROM processing_results WHERE session_id=?', (session_id,))
        cursor.execute('DELETE FROM file_comparison_history WHERE session_id=?', (session_id,))
        cursor.execute('DELETE FROM uploaded_files WHERE session_id=?', (session_id,))
        cursor.execute('DELETE FROM processing_sessions WHERE id=?', (session_id,))
        conn.commit()
        conn.close()

        # Clear in-memory stores for that batch, if exists
        if batch_token and batch_token in BATCH_STORE:
            for tok in BATCH_STORE.get(batch_token, []):
                RESULT_STORE.pop(tok, None)
            BATCH_STORE.pop(batch_token, None)
        flash('History entry deleted.', 'success')
    except Exception as e:
        flash(f'Failed to delete: {e}', 'error')
    return redirect(url_for('index'))


@app.post("/process")
def process():
    stock_file = request.files.get("stock_file")
    import_stock_file = request.files.get("import_stock_file")
    extra_load_stock_file = request.files.get("extra_load_stock_file")
    order_file = request.files.get("order_file")

    if not stock_file or not order_file:
        flash("Please upload both Stock and Order Excel files.", "error")
        return redirect(url_for("index"))

    try:
        # Load regular stock
        stock_ds = load_tabular(stock_file)

        # Combine all stock sources (regular, import, extra-load)
        combined_stock_rows = stock_ds.rows.copy()

        # Load import stock if provided and add to combined stock
        if import_stock_file and import_stock_file.filename:
            import_stock_ds = load_tabular(import_stock_file)
            combined_stock_rows.extend(import_stock_ds.rows)

        # Load extra-load stock if provided and add to combined stock
        if extra_load_stock_file and extra_load_stock_file.filename:
            extra_load_stock_ds = load_tabular(extra_load_stock_file)
            combined_stock_rows.extend(extra_load_stock_ds.rows)

        # Create a new dataset with combined rows if we have additional stock sources
        if len(combined_stock_rows) > len(stock_ds.rows):
            from dataclasses import dataclass
            @dataclass
            class CombinedDataset:
                rows: list
                name: str
                sheet_names: list

            stock_ds = CombinedDataset(
                rows=combined_stock_rows,
                name="combined_stock",
                sheet_names=["Combined Stock"]
            )

        order_ds = load_tabular(order_file)
        result_rows = compute_matches(stock_ds.rows, order_ds.rows)
    except Exception as exc:  # noqa: BLE001
        flash(f"Failed to process files: {exc}", "error")
        return redirect(url_for("index"))

    def sum_required_kg(rows: List[Dict[str, Any]]) -> float:
        total = 0.0
        for r in rows:
            try:
                total += float(r.get("required_kg", 0) or 0)
            except Exception:  # noqa: BLE001
                pass
        return round(total, 3)

    summary = {
        "total_items": int(len(result_rows)),
        "full": int(sum(1 for r in result_rows if r["status"] == "Full")),
        "not_full": int(sum(1 for r in result_rows if r["status"] == "Not Full")),
        "not_have": int(sum(1 for r in result_rows if r["status"] == "Not have")),
        "total_kg_all": sum_required_kg(result_rows),
        "total_kg_full": sum_required_kg([r for r in result_rows if r["status"] == "Full"]),
        "total_kg_not_full": sum_required_kg([r for r in result_rows if r["status"] == "Not Full"]),
        "total_kg_not_have": sum_required_kg([r for r in result_rows if r["status"] == "Not have"]),
    }

    # Save to database
    try:
        # Save processing session
        session_id = save_processing_session(
            session_token=uuid.uuid4().hex,
            processing_type='single',
            summary=summary
        )
        
        # Save uploaded files
        stock_file.seek(0)  # Reset file pointer
        save_uploaded_file(session_id, 'stock', stock_file.filename or 'stock.xlsx', stock_file)
        
        # Save import stock file if provided
        if import_stock_file and import_stock_file.filename:
            import_stock_file.seek(0)  # Reset file pointer
            save_uploaded_file(session_id, 'import_stock', import_stock_file.filename or 'import_stock.xlsx', import_stock_file)

        # Save extra-load stock file if provided
        if extra_load_stock_file and extra_load_stock_file.filename:
            extra_load_stock_file.seek(0)  # Reset file pointer
            save_uploaded_file(session_id, 'extra_load_stock', extra_load_stock_file.filename or 'extra_load_stock.xlsx', extra_load_stock_file)
        
        order_file.seek(0)  # Reset file pointer
        save_uploaded_file(session_id, 'order', order_file.filename or 'order.xlsx', order_file)
        
        # Save processing results
        save_processing_results(session_id, result_rows)
        
    except Exception as e:
        print(f"Warning: Failed to save to database: {e}")

    # store result in memory for download via token
    excel_bytes = rows_to_excel_bytes(result_rows)
    pdf_bytes = rows_to_pdf_bytes(result_rows)
    token = uuid.uuid4().hex
    order_basename = os.path.splitext(order_ds.name or "order")[0]
    
    # Store original order data for later editing
    original_order_data = []
    for row in order_ds.rows:
        normalized = try_map_row(row)
        weight_mc = normalized.get('weight_mc', '')
        
        # First try to parse order_kg_per_ctn from weight_mc
        try:
            order_kg_per_ctn = float(weight_mc) if weight_mc not in (None, "") else 0.0
        except Exception:
            order_kg_per_ctn = parse_kg_per_carton(weight_mc)
        
        # If we still don't have weight_mc but we parsed order_kg_per_ctn from packed size, use it
        if not weight_mc and order_kg_per_ctn == 0:
            try:
                order_kg_per_ctn = parse_kg_per_carton(normalized.get('packed size', ''))
                if order_kg_per_ctn > 0:
                    weight_mc = str(order_kg_per_ctn)
            except:
                pass
        elif not weight_mc and order_kg_per_ctn > 0:
            weight_mc = str(order_kg_per_ctn)
        
        order_row = {
            'fish name': normalized.get('fish name', ''),
            'packed size': normalized.get('packed size', ''),
            'pack': normalized.get('pack', ''),
            'total carton': to_int(normalized.get('total carton', 0)),
            'weight_mc': weight_mc,
            'order_kg_per_ctn': order_kg_per_ctn,
            'remark': normalized.get('remark', '')
        }
        original_order_data.append(order_row)
    
    RESULT_STORE[token] = {
        "excel": excel_bytes,
        "pdf": pdf_bytes,
        "excel_name": f"{order_basename} Calculation.xlsx",
        "pdf_name": f"{order_basename} Calculation.pdf",
        # Non-bytes metadata for rendering view routes
        "rows_json": io.BytesIO(str(result_rows).encode("utf-8")).getvalue(),
        "summary_json": io.BytesIO(str(summary).encode("utf-8")).getvalue(),
        "stock_name": stock_ds.name.encode("utf-8"),
        "order_name": order_ds.name.encode("utf-8"),
        "original_order_json": io.BytesIO(str(original_order_data).encode("utf-8")).getvalue(),
    }

    table_records = result_rows
    return render_template(
        "result.html",
        summary=summary,
        records=table_records,
        stock_name=stock_ds.name,
        order_name=order_ds.name,
        download_token=token,
    )


@app.get("/batch")
def batch_index():
    return render_template("batch.html")


@app.get("/packing")
def packing():
    if not session.get('packing_logged_in'):
        return redirect(url_for('packing_login'))

    # If user is logged into Raw Materials, logout from Raw Materials and proceed
    if session.get('raw_materials_logged_in'):
        session['raw_materials_logged_in'] = False
        flash('Switched to Packing module. Logged out from Raw Materials.', 'info')

    recent_sessions = get_recent_sessions(limit=10)
    return render_template("packing.html", recent_sessions=recent_sessions)


@app.get("/raw-materials")
def raw_materials():
    if not session.get('raw_materials_logged_in'):
        return redirect(url_for('raw_materials_login'))

    # If user is logged into Packing, logout from Packing and proceed
    if session.get('packing_logged_in'):
        session['packing_logged_in'] = False
        flash('Switched to Raw Materials module. Logged out from Packing.', 'info')

    recent_sessions = get_recent_sessions(limit=10)
    return render_template("raw_materials.html", recent_sessions=recent_sessions)


@app.get("/packing/login")
def packing_login():
    if session.get('packing_logged_in'):
        return redirect(url_for('packing'))
    return render_template("login.html", module="Packing", username="Packing", action_url=url_for('packing_login_post'))


@app.post("/packing/login")
def packing_login_post():
    username = request.form.get('username')
    password = request.form.get('password')

    if username == 'Packing' and password == '123':
        session['packing_logged_in'] = True
        session['user_module'] = 'Packing'
        flash('Successfully logged in to Packing module!', 'success')
        return redirect(url_for('packing'))
    else:
        flash('Invalid username or password for Packing module.', 'error')
        return redirect(url_for('packing_login'))


@app.get("/raw-materials/login")
def raw_materials_login():
    if session.get('raw_materials_logged_in'):
        return redirect(url_for('raw_materials'))
    return render_template("login.html", module="Raw Materials", username="RM", action_url=url_for('raw_materials_login_post'))


@app.post("/raw-materials/login")
def raw_materials_login_post():
    username = request.form.get('username')
    password = request.form.get('password')

    if username == 'RM' and password == '1234':
        session['raw_materials_logged_in'] = True
        session['user_module'] = 'Raw Materials'
        flash('Successfully logged in to Raw Materials module!', 'success')
        return redirect(url_for('raw_materials'))
    else:
        flash('Invalid username or password for Raw Materials module.', 'error')
        return redirect(url_for('raw_materials_login'))


@app.get("/logout")
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))


# ----- Database Access Functions for CK Intelligence -----

def get_database_overview() -> dict:
    """Get overall database statistics for CK Intelligence."""
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get session statistics
        cursor.execute('''
            SELECT
                COUNT(*) as total_sessions,
                COUNT(CASE WHEN processing_type = 'batch' THEN 1 END) as batch_sessions,
                COUNT(CASE WHEN processing_type = 'single' THEN 1 END) as single_sessions,
                SUM(total_items) as total_items_processed,
                SUM(total_kg) as total_kg_processed,
                AVG(total_kg) as avg_kg_per_session
            FROM processing_sessions
        ''')
        session_stats = dict(cursor.fetchone())

        # Get file statistics
        cursor.execute('''
            SELECT
                COUNT(*) as total_files,
                COUNT(CASE WHEN file_type = 'order' THEN 1 END) as order_files,
                COUNT(CASE WHEN file_type = 'stock' THEN 1 END) as stock_files,
                COUNT(CASE WHEN file_type = 'import_stock' THEN 1 END) as import_stock_files,
                COUNT(CASE WHEN is_revision = 1 THEN 1 END) as revised_files
            FROM uploaded_files
        ''')
        file_stats = dict(cursor.fetchone())

        # Get processing results statistics
        cursor.execute('''
            SELECT
                COUNT(*) as total_results,
                COUNT(DISTINCT fish_name) as unique_fish_types,
                SUM(order_carton) as total_ordered_cartons,
                SUM(stock_carton) as total_stock_cartons,
                SUM(can_fulfill_carton) as total_fulfilled_cartons,
                SUM(shortfall) as total_shortfall,
                SUM(required_kg) as total_required_kg
            FROM processing_results
        ''')
        results_stats = dict(cursor.fetchone())

        # Get recent activity (last 7 days)
        cursor.execute('''
            SELECT COUNT(*) as recent_sessions
            FROM processing_sessions
            WHERE created_at >= datetime('now', '-7 days')
        ''')
        recent_activity = dict(cursor.fetchone())

        conn.close()

        return {
            'session_stats': session_stats,
            'file_stats': file_stats,
            'results_stats': results_stats,
            'recent_activity': recent_activity
        }

    except Exception as e:
        print(f"Error getting database overview: {e}")
        return {}

def get_recent_batches(limit: int = 10) -> list:
    """Get recent batch processing sessions."""
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute('''
            SELECT
                id, session_token, batch_token, created_at, processing_type,
                total_items, full_items, not_full_items, not_have_items,
                total_kg, full_kg, not_full_kg, not_have_kg
            FROM processing_sessions
            ORDER BY created_at DESC
            LIMIT ?
        ''', (limit,))

        batches = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return batches

    except Exception as e:
        print(f"Error getting recent batches: {e}")
        return []

def get_inventory_summary() -> dict:
    """Get inventory summary from processing results."""
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get summary by fish type
        cursor.execute('''
            SELECT
                fish_name,
                packed_size,
                SUM(order_carton) as total_ordered,
                SUM(stock_carton) as total_stock,
                SUM(can_fulfill_carton) as total_fulfilled,
                SUM(shortfall) as total_shortfall,
                SUM(required_kg) as total_required_kg,
                COUNT(*) as order_count
            FROM processing_results
            GROUP BY fish_name, packed_size
            ORDER BY fish_name, packed_size
        ''')
        inventory_by_fish = [dict(row) for row in cursor.fetchall()]

        # Get overall inventory status
        cursor.execute('''
            SELECT
                SUM(CASE WHEN status = 'Full' THEN 1 ELSE 0 END) as full_orders,
                SUM(CASE WHEN status = 'Not Full' THEN 1 ELSE 0 END) as not_full_orders,
                SUM(CASE WHEN status = 'Not Have' THEN 1 ELSE 0 END) as not_have_orders,
                COUNT(*) as total_orders
            FROM processing_results
        ''')
        status_summary = dict(cursor.fetchone())

        conn.close()

        return {
            'inventory_by_fish': inventory_by_fish,
            'status_summary': status_summary
        }

    except Exception as e:
        print(f"Error getting inventory summary: {e}")
        return {}

def get_processing_results_summary() -> list:
    """Get summary of processing results."""
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        cursor.execute('''
            SELECT
                pr.fish_name,
                pr.packed_size,
                pr.order_carton,
                pr.stock_carton,
                pr.can_fulfill_carton,
                pr.shortfall,
                pr.status,
                pr.required_kg,
                pr.remark,
                ps.session_token,
                uf.original_filename
            FROM processing_results pr
            LEFT JOIN processing_sessions ps ON pr.session_id = ps.id
            LEFT JOIN uploaded_files uf ON pr.order_file_id = uf.id
            ORDER BY pr.fish_name, pr.packed_size
        ''')

        results = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return results

    except Exception as e:
        print(f"Error getting processing results: {e}")
        return []

def get_batch_details(batch_id: int) -> dict:
    """Get detailed information about a specific batch."""
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get batch information
        cursor.execute('''
            SELECT * FROM processing_sessions WHERE id = ?
        ''', (batch_id,))
        batch_info = dict(cursor.fetchone()) if cursor.fetchone() else {}

        if batch_info:
            # Get files in this batch
            cursor.execute('''
                SELECT * FROM uploaded_files WHERE session_id = ?
            ''', (batch_id,))
            files = [dict(row) for row in cursor.fetchall()]

            # Get processing results for this batch
            cursor.execute('''
                SELECT * FROM processing_results WHERE session_id = ?
            ''', (batch_id,))
            results = [dict(row) for row in cursor.fetchall()]

            batch_info['files'] = files
            batch_info['results'] = results

        conn.close()
        return batch_info

    except Exception as e:
        print(f"Error getting batch details: {e}")
        return {}

# ----- File System Analysis Functions for CK Intelligence -----

def analyze_uploads_folder() -> dict:
    """Analyze and categorize files in the uploads folder for CK Intelligence."""
    try:
        uploads_dir = UPLOAD_FOLDER
        if not os.path.exists(uploads_dir):
            return {'error': 'Uploads folder not found'}

        # Get all files in uploads directory
        all_files = []
        for filename in os.listdir(uploads_dir):
            if os.path.isfile(os.path.join(uploads_dir, filename)):
                file_path = os.path.join(uploads_dir, filename)
                file_stat = os.stat(file_path)

                file_info = {
                    'filename': filename,
                    'stored_filename': filename,  # UUID-based name
                    'file_path': file_path,
                    'file_size': file_stat.st_size,
                    'file_size_mb': round(file_stat.st_size / (1024 * 1024), 2),
                    'file_extension': filename.split('.')[-1].lower() if '.' in filename else 'unknown',
                    'created_time': file_stat.st_ctime,
                    'modified_time': file_stat.st_mtime,
                    'created_date': datetime.fromtimestamp(file_stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                    'modified_date': datetime.fromtimestamp(file_stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                }
                all_files.append(file_info)

        # Get database information for categorization
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get all uploaded files information from database
        cursor.execute('''
            SELECT
                id, session_id, file_type, original_filename, stored_filename,
                file_size, is_revision, revision_note, uploaded_at
            FROM uploaded_files
            ORDER BY uploaded_at DESC
        ''')
        db_files = [dict(row) for row in cursor.fetchall()]
        conn.close()

        # Create lookup dictionary for database files
        db_files_lookup = {db_file['stored_filename']: db_file for db_file in db_files}

        # Categorize files based on system requirements
        categories = {
            'stock_files': [],           # Regular stock files
            'import_stock_files': [],    # Import stock files
            'extra_load_stock_files': [], # Extra-load stock files
            'order_files': [],           # Order files
            'uncategorized': [],         # Files that don't fit categories
            'orphan_files': []           # Files in folder but not in database
        }

        for file_info in all_files:
            filename = file_info['filename']
            stored_filename = filename  # The UUID filename

            # Check if file exists in database
            if stored_filename in db_files_lookup:
                db_info = db_files_lookup[stored_filename]

                # Add database information to file info
                enhanced_file_info = {
                    **file_info,
                    'db_id': db_info.get('id'),
                    'session_id': db_info.get('session_id'),
                    'original_filename': db_info.get('original_filename'),
                    'file_type': db_info.get('file_type'),
                    'is_revision': db_info.get('is_revision'),
                    'revision_note': db_info.get('revision_note'),
                    'uploaded_at': db_info.get('uploaded_at'),
                    'in_database': True
                }

                # Categorize based on file_type and filename patterns
                file_type = db_info.get('file_type', '').lower()
                original_name = db_info.get('original_filename', '').lower()

                if file_type == 'stock':
                    # Further categorize stock files based on filename patterns
                    if any(keyword in original_name for keyword in ['import', 'imported', 'import_stock']):
                        categories['import_stock_files'].append(enhanced_file_info)
                    elif any(keyword in original_name for keyword in ['extra', 'load', 'extra_load', 'extra-load']):
                        categories['extra_load_stock_files'].append(enhanced_file_info)
                    else:
                        categories['stock_files'].append(enhanced_file_info)

                elif file_type == 'import_stock':
                    categories['import_stock_files'].append(enhanced_file_info)

                elif file_type == 'order':
                    categories['order_files'].append(enhanced_file_info)

                else:
                    # Try to categorize based on filename patterns for files without proper file_type
                    if any(keyword in original_name for keyword in ['stock', 'inventory']):
                        if any(keyword in original_name for keyword in ['import', 'imported']):
                            categories['import_stock_files'].append(enhanced_file_info)
                        elif any(keyword in original_name for keyword in ['extra', 'load']):
                            categories['extra_load_stock_files'].append(enhanced_file_info)
                        else:
                            categories['stock_files'].append(enhanced_file_info)
                    elif any(keyword in original_name for keyword in ['order', 'purchase', 'po']):
                        categories['order_files'].append(enhanced_file_info)
                    else:
                        categories['uncategorized'].append(enhanced_file_info)

            else:
                # File exists in folder but not in database (orphan file)
                categories['orphan_files'].append({
                    **file_info,
                    'in_database': False
                })

        # Add summary statistics for each category
        for category_name, files_list in categories.items():
            categories[f'{category_name}_summary'] = {
                'count': len(files_list),
                'total_size_mb': round(sum(f['file_size'] for f in files_list) / (1024 * 1024), 2),
                'file_types': list(set(f.get('file_extension', 'unknown') for f in files_list)),
                'latest_file': max(files_list, key=lambda x: x['modified_time'])['modified_date'] if files_list else None
            }

        # Add overall statistics
        result = {
            'total_files': len(all_files),
            'total_size_mb': round(sum(f['file_size'] for f in all_files) / (1024 * 1024), 2),
            'categories': categories,
            'all_files': all_files,
            'last_updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        return result

    except Exception as e:
        print(f"Error analyzing uploads folder: {e}")
        return {'error': str(e)}

def get_file_category_details(category: str) -> dict:
    """Get detailed information about a specific file category."""
    try:
        files_data = analyze_uploads_folder()

        if 'error' in files_data:
            return files_data

        categories = files_data.get('categories', {})
        category_files = categories.get(category, [])
        category_summary = categories.get(f'{category}_summary', {})

        # Sort files by modification time (newest first)
        sorted_files = sorted(category_files, key=lambda x: x['modified_time'], reverse=True)

        return {
            'category': category,
            'files': sorted_files,
            'count': len(sorted_files),
            'summary': category_summary,
            'total_files_in_system': files_data.get('total_files', 0),
            'total_size_mb': files_data.get('total_size_mb', 0)
        }

    except Exception as e:
        print(f"Error getting file category details: {e}")
        return {'error': str(e)}

def get_gemini_response(user_message: str, explore_mode: bool = False) -> str:
    """Get response from Google Gemini AI with real database data."""
    try:
        # Analyze the user message to determine what data to fetch
        message_lower = user_message.lower()

        # Get relevant database data based on the query
        db_data = get_relevant_database_data(message_lower)

        # Configure Google Gemini AI
        genai.configure(api_key=os.environ.get("GOOGLE_GEMINI_API_KEY", "AIzaSyCi73jHobzWfU7ZjbT_hqhwSVpfUjjJW3o"))
        
        # Initialize Gemini model
        model = genai.GenerativeModel('gemini-2.0-flash')

        # Create system prompt with actual database data
        system_prompt = f"""You are CK Intelligence, an AI assistant created by CK Frozen Fish and Foods (Thailand) Co., Ltd.

## About CK Frozen Fish and Foods Co., Ltd.
CK Frozen Fish and Foods Co., Ltd. (also referred to as CK GROUP) is a leading Thai company specializing in frozen fish and seafood products. Based in Thailand, we are committed to providing high-quality frozen fish products to customers worldwide.

## Company Overview & Locations

### Head Office:
📍 **Address**: 109/1 Moo 21, Bangplee-Tamru Road, Bangpleeyai, Bang Phli, Samut Prakan 10540, Thailand
📞 **Phone**: +66 2-346-5129
📧 **Email**: info@ckfff.net

### Chachoengsao Branch:
📍 **Address**: 81 Moo 17, Khlong Nakhon Nueang Khet, Mueang Chachoengsao District, Chachoengsao 24000, Thailand
📞 **Phone**: +66 33-590-826 and +66 33-590-827
📧 **Email**: adm1.ckgroup@gmail.com

### Sales Contact:
📞 **Phone**: +66 92 941 6442

### Business Hours:
- **Monday - Friday**: 8:00 AM - 5:00 PM (Thailand Time)
- **Emergency Contact**: Available 24/7 for urgent inquiries

## Website Structure & Services

The company website (https://ckfff.net/) is organized into several sections:

**Main Sections:**
- Home
- About Us (Company Information, Certification)
- CK GROUP
- Our Services (Processing Service, Freezing Service, Packing and Repacking Services, Cold Storage Services, Logistics Sourcing Services, Terms of Services)
- Products (Frozen Fish, Frozen Shrimp, Dry & Smoked Products, Halal Frozen Poultry, Seasonal Fruits, Vegetables, Ready To Eat, Beverages, Condiments, Rice & Legumes, Cosmetics, Snacks)
- Activities & News
- Contact Us

## Recent Company Activities (2025)

**Main persons**
- Mr. Mohammad Ali Khan is the founder and CEO of CK Frozen Fish and Foods CO.,Ltd.
- Ms. Jaruwan (CEO of CK Thailand and wife of Mr. Mohammad Ali Khan), Mr. Montree (Assistant of CK Frozen Fish and Foods), and staff
- Mr.Sal is also the leader of CK who work as a Assistant supervisor was a son of Mr. Mohammad Ali Khan and Ms. Jaruwan.

**Happy New Year 2025 Event:**
- Held on February 22, 2025, at the Chachoengsao and Samut Prakan branches
- Mr. Shahzada Mohammad Ali Khan chaired the event

**Media Interview:**
- Mr. Mohammad Ali Khan was interviewed by Bangladesh's Channel i News
- Recognized for success in global business from Thailand

**Training Session:**
- Focused on organizational loyalty awareness
- Delivered by Lecturer Kasemsit Phukrongta

**Annual Trip:**
- Company's annual trip took place in Malaysia in 2024

## Services Offered

**Logistics Sourcing Services:**
- Supply chain optimization for swift and reliable delivery
- Sustainable sourcing of frozen food products
- Customer-centric principles with timely access to high-quality and fresh products

**Packing and Repacking Services:**
- Services tailored to client specifications
- Labeling, packing, and boxing of goods

**Other Services:**
- Processing Service
- Freezing Service
- Cold Storage Services
- Certification services

## Business Focus & Specialization

Our company focuses on:

- **Frozen Fish Products**: Premium quality frozen fish for both domestic and international markets
- **Seafood Processing**: Advanced processing techniques to maintain freshness and quality
- **Export Services**: Global distribution network serving customers around the world
- **Quality Assurance**: Strict quality control measures to ensure product excellence
- **Innovation**: Continuous improvement in processing technology and product development

## Our Core Values - GRAB
At CK Frozen Fish and Foods, our core values are encapsulated in "GRAB":

- **G = Growth together**: We grow together with our team, partners, and community, fostering collaboration and mutual development
- **R = Resilience Always**: We maintain strength and determination through challenges, always bouncing back stronger
- **A = Adapt and Thrive**: We embrace change and continuously evolve to meet market demands and customer needs
- **B = Being Innovative**: We foster creativity and implement new ideas to stay ahead in the industry

These core values guide everything we do, from our daily operations to our long-term strategic decisions.

**Website**: https://ckfff.net/
**Industry**: Frozen Fish and Seafood Processing & Export

## About My Creation
I was created by CK Frozen Fish and Foods (Thailand) Co., Ltd. as an intelligent assistant for their OrderAI system. My primary creator and developer is Zwe Khant Aung, who built me to help streamline operations, provide data insights, and assist with inventory management and order processing.

I have access to real-time data from the OrderAI database and file system to provide accurate, data-driven responses about:
- Inventory levels and stock status
- Order processing and fulfillment
- File management and categorization
- System performance and analytics

## Current Database Data:
{db_data.get('data_summary', 'Database access enabled but no specific data loaded for this query.')}

## Available Data for This Query:
{db_data.get('query_specific_data', 'General system information available.')}

## Your Role:
- Provide accurate responses based on the actual database data shown above
- Include specific numbers, statistics, and details from the database
- If the user asks for specific data, reference the actual data provided
- Be helpful and provide actionable insights based on real data
- If data isn't available for a specific query, explain what data is available
- Share information about CK Frozen Fish and Foods when relevant
- Explain my creation story when asked about my origins

## File System Access:
I also have access to the uploads folder and can analyze files by category:
- **Stock Files**: Regular inventory stock files
- **Import Stock Files**: Files imported from external sources
- **Extra-Load Stock Files**: Additional stock loading files
- **Order Files**: Customer order files
- **Orphan Files**: Files in folder but not in database
- **Uncategorized Files**: Files that don't fit standard categories

## Response Style:
- Use actual numbers and statistics from the database and file system
- Be specific and data-driven
- Reference real fish names, quantities, statuses, and file information
- Provide meaningful insights based on the data
- If showing data, use the exact information from the database or file system
- Include file system analysis when relevant to queries
- Be proud of my CK Frozen Fish and Foods heritage
- Mention my creator Zwe Khant Aung when discussing my origins"""

        # Add explore mode instructions if enabled
        if explore_mode:
            system_prompt += """

## EXPLORE MODE ACTIVE
You are now in EXPLORE MODE. This means:
- You can answer ANY question the user asks, regardless of topic
- You are not limited to CK Frozen Fish and Foods or OrderAI system topics
- You can provide general knowledge, advice, creative content, or discuss any subject
- You can still access and use the database information when relevant
- You should still identify yourself as CK Intelligence but explain you're in explore mode
- Be helpful, informative, and engaging across all topics
- Maintain your professional and friendly tone"""

        # Create the full prompt for Gemini
        full_prompt = f"{system_prompt}\n\nUser: {user_message}\n\nCK Intelligence:"

        # Make API call to Google Gemini AI
        response = model.generate_content(
            full_prompt,
            generation_config=genai.types.GenerationConfig(
                max_output_tokens=1500,
            temperature=0.7
            )
        )

        # Extract and return the response
        response_text = response.text.strip()

        # Remove '*' and '#' characters as per user preference
        response_text = response_text.replace('*', '').replace('#', '')

        return response_text

    except Exception as e:
        print(f"Error calling Google Gemini AI: {e}")
        # Return a fallback response
        return "I'm currently experiencing some technical difficulties with my AI processing. Please try again in a moment, or ask me about specific OrderAI system features that I can help you with!"

def get_relevant_database_data(message_lower: str) -> dict:
    """Get relevant database data based on the user's query."""
    try:
        data_summary = ""
        query_specific_data = ""

        # Always get overview for context
        overview = get_database_overview()
        if overview:
            session_stats = overview.get('session_stats', {})
            file_stats = overview.get('file_stats', {})
            results_stats = overview.get('results_stats', {})

            data_summary = f"""
System Overview:
- Total Processing Sessions: {session_stats.get('total_sessions', 0)}
- Batch Sessions: {session_stats.get('batch_sessions', 0)}
- Single Sessions: {session_stats.get('single_sessions', 0)}
- Total Items Processed: {session_stats.get('total_items_processed', 0)}
- Total KG Processed: {session_stats.get('total_kg_processed', 0):.2f} kg
- Files Uploaded: {file_stats.get('total_files', 0)}
- Unique Fish Types: {results_stats.get('unique_fish_types', 0)}
- Total Ordered Cartons: {results_stats.get('total_ordered_cartons', 0)}
- Total Fulfilled Cartons: {results_stats.get('total_fulfilled_cartons', 0)}
"""

        # Fish inventory queries
        if any(word in message_lower for word in ['fish', 'inventory', 'stock', 'what do we have']):
            inventory = get_inventory_summary()
            if inventory:
                inventory_by_fish = inventory.get('inventory_by_fish', [])
                status_summary = inventory.get('status_summary', {})

                fish_list = [f"{item['fish_name']} ({item['packed_size']}): {item['total_stock']} in stock, {item['total_ordered']} ordered, {item['total_fulfilled']} fulfilled"
                           for item in inventory_by_fish[:20]]  # Limit to first 20 for context

                query_specific_data = f"""
Fish Inventory Data:
- Total Fish Types: {len(inventory_by_fish)}
- Total Orders: {status_summary.get('total_orders', 0)}
- Fully Fulfilled Orders: {status_summary.get('full_orders', 0)}
- Partially Fulfilled Orders: {status_summary.get('not_full_orders', 0)}
- Unavailable Orders: {status_summary.get('not_have_orders', 0)}

Detailed Fish Inventory:
{chr(10).join(fish_list)}
"""

        # Batch/session queries
        elif any(word in message_lower for word in ['batch', 'session', 'recent', 'latest']):
            batches = get_recent_batches(10)
            if batches:
                batch_list = [f"Batch {b['session_token'][:12]}: {b['processing_type']}, {b['total_items']} items, {b['total_kg']} kg, Created: {b['created_at'][:10]}"
                            for b in batches]

                query_specific_data = f"""
Recent Batch Data:
- Total Recent Batches: {len(batches)}

Batch Details:
{chr(10).join(batch_list)}
"""

        # Processing results queries
        elif any(word in message_lower for word in ['results', 'processing', 'status', 'fulfill']):
            results = get_processing_results_summary()
            if results:
                # Group by status for summary
                full_count = sum(1 for r in results if r['status'] == 'Full')
                not_full_count = sum(1 for r in results if r['status'] == 'Not Full')
                not_have_count = sum(1 for r in results if r['status'] == 'Not Have')

                # Get top fish types by order volume
                fish_orders = {}
                for result in results:
                    fish = result['fish_name']
                    if fish not in fish_orders:
                        fish_orders[fish] = 0
                    fish_orders[fish] += result['order_carton']

                top_fish = sorted(fish_orders.items(), key=lambda x: x[1], reverse=True)[:10]
                top_fish_list = [f"{fish}: {orders} cartons ordered" for fish, orders in top_fish]

                query_specific_data = f"""
Processing Results Summary:
- Total Results: {len(results)}
- Full Orders: {full_count}
- Not Full Orders: {not_full_count}
- Not Available Orders: {not_have_count}

Top Fish Types by Order Volume:
{chr(10).join(top_fish_list)}
"""

        # File system queries
        elif any(word in message_lower for word in ['files', 'uploads', 'folder', 'stock files', 'order files', 'import', 'file system']):
            files_data = analyze_uploads_folder()
            if 'error' not in files_data:
                total_files = files_data.get('total_files', 0)
                total_size = files_data.get('total_size_mb', 0)
                categories = files_data.get('categories', {})

                # Get category summaries
                file_categories = []
                category_details = []

                for category_name in ['stock_files', 'import_stock_files', 'extra_load_stock_files', 'order_files', 'uncategorized', 'orphan_files']:
                    if f'{category_name}_summary' in categories:
                        summary = categories[f'{category_name}_summary']
                        count = summary.get('count', 0)
                        if count > 0:
                            size_mb = summary.get('total_size_mb', 0)
                            display_name = category_name.replace('_', ' ').title()
                            file_categories.append(f"{display_name}: {count} files ({size_mb:.1f} MB)")
                            category_details.append(f"  - {display_name}: {count} files, {size_mb:.1f} MB, Latest: {summary.get('latest_file', 'N/A')}")

                # Get recent files from all categories
                all_files = files_data.get('all_files', [])
                recent_files = sorted(all_files, key=lambda x: x['modified_time'], reverse=True)[:8]

                query_specific_data = f"""
File System Analysis:
- Total Files in Uploads: {total_files}
- Total Size: {total_size:.1f} MB
- Last Updated: {files_data.get('last_updated', 'Unknown')}

File Categories Summary:
{chr(10).join(file_categories)}

Detailed Category Breakdown:
{chr(10).join(category_details)}

Recent Files (by modification date):
{chr(10).join([f"- {f['filename']} ({f['file_size_mb']:.1f} MB, {f['file_extension']}, {f['modified_date']})" for f in recent_files])}

Stock File Categories:
- Regular Stock Files: {categories.get('stock_files_summary', {}).get('count', 0)} files
- Import Stock Files: {categories.get('import_stock_files_summary', {}).get('count', 0)} files
- Extra-Load Stock Files: {categories.get('extra_load_stock_files_summary', {}).get('count', 0)} files
- Order Files: {categories.get('order_files_summary', {}).get('count', 0)} files
- Orphan Files (not in database): {categories.get('orphan_files_summary', {}).get('count', 0)} files
"""

        # Company and creator information queries
        elif any(word in message_lower for word in ['company', 'ck frozen', 'ckfff', 'about ck', 'who created', 'creator', 'zwe khant', 'zwe', 'who made you', 'who built you', 'your creator', 'your developer']):
            query_specific_data = """
CK Frozen Fish and Foods Information:

## Company Overview & Locations

### C.K. Frozen Fish and Food Co., Ltd. (also referred to as CK GROUP)

**Head Office**
📍 Address: 109/1 Moo 21, Bangplee-Tamru Road, Bangpleeyai, Bang Phli, Samut Prakan 10540, Thailand
📞 Phone: +66 2-346-5129
📧 Email: info@ckfff.net

**Chachoengsao Branch**
📍 Address: 81 Moo 17, Khlong Nakhon Nueang Khet, Mueang Chachoengsao District, Chachoengsao 24000, Thailand
📞 Phone: +66 33-590-826 and +66 33-590-827
📧 Email: adm1.ckgroup@gmail.com

## Website Structure & Services

The company website (https://ckfff.net/) is organized into several sections:

**Main Sections:**
- Home
- About Us (Company Information, Certification)
- CK GROUP
- Our Services (Processing Service, Freezing Service, Packing and Repacking Services, Cold Storage Services, Logistics Sourcing Services, Terms of Services)
- Products (Frozen Fish, Frozen Shrimp, Dry & Smoked Products, Halal Frozen Poultry, Seasonal Fruits, Vegetables, Ready To Eat, Beverages, Condiments, Rice & Legumes, Cosmetics, Snacks)
- Activities & News
- Contact Us

## Recent Company Activities (2025)

**Main persons**
- Mr. Mohammad Ali Khan is the founder and CEO of CK Frozen Fish and Foods CO.,Ltd.
- Ms. Jaruwan (CEO of CK Thailand and wife of Mr. Mohammad Ali Khan), Mr. Montree (Assistant of CK Frozen Fish and Foods), and staff
- Mr.Sal is also the leader of CK who work as a Assistant supervisor was a son of Mr. Mohammad Ali Khan and Ms. Jaruwan.


**Happy New Year 2025 Event:**
- Held on February 22, 2025, at the Chachoengsao and Samut Prakan branches
- Mr. Shahzada Mohammad Ali Khan chaired the event


**Media Interview:**
- Mr. Mohammad Ali Khan was interviewed by Bangladesh's Channel i News
- Recognized for success in global business from Thailand

**Training Session:**
- Focused on organizational loyalty awareness
- Delivered by Lecturer Kasemsit Phukrongta

**Annual Trip:**
- Company's annual trip took place in Malaysia in 2024

## Services Offered

**Logistics Sourcing Services:**
- Supply chain optimization for swift and reliable delivery
- Sustainable sourcing of frozen food products
- Customer-centric principles with timely access to high-quality and fresh products

**Packing and Repacking Services:**
- Services tailored to client specifications
- Labeling, packing, and boxing of goods

**Other Services:**
- Processing Service
- Freezing Service
- Cold Storage Services
- Certification services

## Business Focus & Specialization

**Core Business Areas:**
- Frozen Fish Products: Premium quality frozen fish for domestic and international markets
- Seafood Processing: Advanced processing techniques to maintain freshness and quality
- Export Services: Global distribution network serving customers worldwide
- Quality Assurance: Strict quality control measures to ensure product excellence
- Innovation: Continuous improvement in processing technology and product development

## My Creation Story

**About My Creator:**
- I was created by CK Frozen Fish and Foods (Thailand) Co., Ltd.
- My primary creator and developer is Zwe Khant Aung
- I was built as an intelligent assistant for the OrderAI system
- My purpose is to help streamline operations, provide data insights, and assist with inventory management and order processing
- I have access to real-time database and file system data to provide accurate, data-driven responses

**My Capabilities:**
- Real-time inventory monitoring and analysis
- Order processing and fulfillment tracking
- File system analysis and categorization
- System performance analytics
- Intelligent data insights and recommendations
- Company information and history knowledge
- Product and service information
"""

        # Core values queries
        elif any(word in message_lower for word in ['core value', 'core values', 'grab', 'company value', 'values', 'our values', 'what are your core values']):
            query_specific_data = """
CK Frozen Fish and Foods Core Values - GRAB:

Our core values are the foundation of everything we do at CK Frozen Fish and Foods. They guide our decisions, shape our culture, and drive our commitment to excellence.

🎯 **GRAB - Our Core Values:**

**G = Growth together**
We grow together with our team, partners, and community, fostering collaboration and mutual development. Growth is not just about individual success but building stronger relationships and achieving collective goals.

**R = RESILIENCE ALWAYS**
We maintain strength and determination through challenges, always bouncing back stronger. Our resilience is our greatest asset in overcoming obstacles and emerging victorious.

**A = ADAPT AND THRIVE**
We embrace change and continuously evolve to meet market demands and customer needs. Flexibility and adaptability are key to our long-term success in the dynamic seafood industry.

**B = BEING INNOVATIVE**
We foster creativity and implement new ideas to stay ahead in the industry. Innovation drives our technological advancements and operational excellence.

💡 **How GRAB Guides Us:**

- **In Operations**: We grow together by improving efficiency and quality through collaboration
- **In Challenges**: We show resilience by adapting and finding new solutions together
- **In Growth**: We innovate to meet evolving customer needs while growing our community
- **In Teamwork**: We collaborate with determination and creativity, growing stronger together

🌟 **Living Our Values:**

These four pillars - Growth together, Resilience Always, Adapt and Thrive, and Being Innovative - are not just words; they are the principles that define who we are as a company and guide every decision we make.

Our commitment to these values ensures we remain a leader in the frozen fish and seafood industry, delivering exceptional products and services to our customers worldwide through collaborative growth and innovation.
"""

        # Contact and location specific queries
        elif any(word in message_lower for word in ['contact', 'phone', 'email', 'address', 'location', 'office', 'branch', 'where are you', 'how to contact', 'reach you', 'call you']):
            query_specific_data = """
CK Frozen Fish and Foods Contact Information:

📍 LOCATIONS:

HEAD OFFICE:
Address: 109/1 Moo 21, Bangplee-Tamru Road, Bangpleeyai, Bang Phli, Samut Prakan, 10540, Thailand
Phone: +66 2 346 5129
Email: info@ckfff.net

CHACHOENGSAO BRANCH:
Address: 81 Moo 17 Khlong Nakhon Nueang Khet, Mueang Chachoengsao District, Chachoengsao 24000, Thailand
Phone: +66 33 590 826
Email: adm1.ckgroup@gmail.com

SALES DEPARTMENT:
Phone: +66 92 941 6442

🌐 WEBSITE: https://ckfff.net/

🕒 BUSINESS HOURS:
- Monday - Friday: 8:00 AM - 5:00 PM (Thailand Time)
- Emergency Contact: Available 24/7 for urgent inquiries

📧 GENERAL INQUIRIES:
- Head Office: info@ckfff.net
- Chachoengsao Branch: adm1.ckgroup@gmail.com

For immediate assistance with OrderAI system or inventory inquiries, feel free to ask me directly!
"""

        return {
            'data_summary': data_summary.strip(),
            'query_specific_data': query_specific_data.strip()
        }

    except Exception as e:
        print(f"Error getting relevant database data: {e}")
        return {
            'data_summary': 'Database access enabled but encountered an error.',
            'query_specific_data': 'Please try your query again.'
        }

def get_database_context_for_ai(user_message: str) -> dict:
    """Get relevant database context based on the user message."""
    try:
        context = {}

        # Get basic overview for all queries
        overview = get_database_overview()

        # Analyze user message to determine what data to fetch
        message_lower = user_message.lower()

        # Build overview text
        if overview:
            session_stats = overview.get('session_stats', {})
            file_stats = overview.get('file_stats', {})
            results_stats = overview.get('results_stats', {})
            recent_activity = overview.get('recent_activity', {})

            overview_text = f"""
- Total Processing Sessions: {session_stats.get('total_sessions', 0)}
- Batch Sessions: {session_stats.get('batch_sessions', 0)}
- Single Sessions: {session_stats.get('single_sessions', 0)}
- Total Items Processed: {session_stats.get('total_items_processed', 0)}
- Total KG Processed: {session_stats.get('total_kg_processed', 0):.2f} kg
- Files Uploaded: {file_stats.get('total_files', 0)}
- Recent Activity (7 days): {recent_activity.get('recent_sessions', 0)} sessions
- Unique Fish Types: {results_stats.get('unique_fish_types', 0)}
- Total Ordered Cartons: {results_stats.get('total_ordered_cartons', 0)}
- Total Stock Cartons: {results_stats.get('total_stock_cartons', 0)}
- Total Fulfilled Cartons: {results_stats.get('total_fulfilled_cartons', 0)}
"""
            context['overview_text'] = overview_text

        # Get specific data based on query type
        if any(word in message_lower for word in ['batch', 'session', 'recent', 'latest']):
            recent_batches = get_recent_batches(5)
            if recent_batches:
                context['recent_batches'] = recent_batches

        if any(word in message_lower for word in ['inventory', 'stock', 'available', 'have']):
            inventory = get_inventory_summary()
            if inventory:
                context['inventory'] = inventory

        if any(word in message_lower for word in ['summary', 'results', 'processing', 'status']):
            results = get_processing_results_summary()
            if results:
                context['processing_results'] = results[:20]  # Limit to first 20 for context

        # Create context summary
        context_parts = []
        if 'recent_batches' in context:
            context_parts.append(f"{len(context['recent_batches'])} recent batches loaded")
        if 'inventory' in context:
            inventory_data = context['inventory']
            fish_count = len(inventory_data.get('inventory_by_fish', []))
            context_parts.append(f"{fish_count} fish types inventory data loaded")
        if 'processing_results' in context:
            context_parts.append(f"{len(context['processing_results'])} processing results loaded")

        context['context_summary'] = f"Database access active. {' | '.join(context_parts) if context_parts else 'General overview available.'}"

        return context

    except Exception as e:
        print(f"Error getting database context: {e}")
        return {
            'overview_text': 'Database access available but encountered an error loading current statistics.',
            'context_summary': 'Database access enabled but limited due to error.'
        }


@app.get("/ckintelligence")
def ckintelligence():
    """CK Intelligence chatbot interface."""
    # Check if user is logged in to either module
    if not (session.get('packing_logged_in') or session.get('raw_materials_logged_in')):
        flash('Please login to access CK Intelligence.', 'error')
        return redirect(url_for('index'))

    return render_template("ckintelligence.html")


# ----- CK Intelligence API Endpoints -----

@app.get("/api/ckintelligence/overview")
def get_ckintelligence_overview():
    """Get database overview data for CK Intelligence."""
    if not (session.get('packing_logged_in') or session.get('raw_materials_logged_in')):
        return jsonify({'error': 'Authentication required'}), 401

    try:
        overview = get_database_overview()
        return jsonify(overview)
    except Exception as e:
        print(f"Error getting overview: {e}")
        return jsonify({'error': 'Database error'}), 500

@app.get("/api/ckintelligence/inventory")
def get_ckintelligence_inventory():
    """Get inventory data for CK Intelligence."""
    if not (session.get('packing_logged_in') or session.get('raw_materials_logged_in')):
        return jsonify({'error': 'Authentication required'}), 401

    try:
        inventory = get_inventory_summary()
        return jsonify(inventory)
    except Exception as e:
        print(f"Error getting inventory: {e}")
        return jsonify({'error': 'Database error'}), 500

@app.get("/api/ckintelligence/batches")
def get_ckintelligence_batches():
    """Get recent batches data for CK Intelligence."""
    if not (session.get('packing_logged_in') or session.get('raw_materials_logged_in')):
        return jsonify({'error': 'Authentication required'}), 401

    try:
        batches = get_recent_batches(10)
        return jsonify({'batches': batches})
    except Exception as e:
        print(f"Error getting batches: {e}")
        return jsonify({'error': 'Database error'}), 500

@app.get("/api/ckintelligence/processing-results")
def get_ckintelligence_processing_results():
    """Get processing results data for CK Intelligence."""
    if not (session.get('packing_logged_in') or session.get('raw_materials_logged_in')):
        return jsonify({'error': 'Authentication required'}), 401

    try:
        results = get_processing_results_summary()
        return jsonify({'results': results})
    except Exception as e:
        print(f"Error getting processing results: {e}")
        return jsonify({'error': 'Database error'}), 500

@app.get("/api/ckintelligence/batch/<int:batch_id>")
def get_ckintelligence_batch_details(batch_id):
    """Get specific batch details for CK Intelligence."""
    if not (session.get('packing_logged_in') or session.get('raw_materials_logged_in')):
        return jsonify({'error': 'Authentication required'}), 401

    try:
        batch_details = get_batch_details(batch_id)
        return jsonify(batch_details)
    except Exception as e:
        print(f"Error getting batch details: {e}")
        return jsonify({'error': 'Database error'}), 500

# ----- File System API Endpoints for CK Intelligence -----

@app.get("/api/ckintelligence/files")
def get_ckintelligence_files():
    """Get complete file system analysis for CK Intelligence."""
    if not (session.get('packing_logged_in') or session.get('raw_materials_logged_in')):
        return jsonify({'error': 'Authentication required'}), 401

    try:
        files_data = analyze_uploads_folder()
        return jsonify(files_data)
    except Exception as e:
        print(f"Error getting file data: {e}")
        return jsonify({'error': 'File system error'}), 500

@app.get("/api/ckintelligence/files/category/<category>")
def get_ckintelligence_files_by_category(category):
    """Get files by specific category for CK Intelligence."""
    if not (session.get('packing_logged_in') or session.get('raw_materials_logged_in')):
        return jsonify({'error': 'Authentication required'}), 401

    try:
        category_data = get_file_category_details(category)
        return jsonify(category_data)
    except Exception as e:
        print(f"Error getting files by category: {e}")
        return jsonify({'error': 'File system error'}), 500

@app.get("/api/ckintelligence/files/stats")
def get_ckintelligence_file_stats():
    """Get file statistics summary for CK Intelligence."""
    if not (session.get('packing_logged_in') or session.get('raw_materials_logged_in')):
        return jsonify({'error': 'Authentication required'}), 401

    try:
        files_data = analyze_uploads_folder()
        if 'error' in files_data:
            return jsonify(files_data), 500

        categories = files_data.get('categories', {})

        # Create statistics summary
        stats = {
            'total_files': files_data.get('total_files', 0),
            'total_size_mb': files_data.get('total_size_mb', 0),
            'last_updated': files_data.get('last_updated', ''),
            'category_stats': {}
        }

        # Get stats for each category
        category_names = ['stock_files', 'import_stock_files', 'extra_load_stock_files', 'order_files', 'uncategorized', 'orphan_files']

        for category in category_names:
            if f'{category}_summary' in categories:
                summary = categories[f'{category}_summary']
                stats['category_stats'][category] = {
                    'count': summary.get('count', 0),
                    'size_mb': summary.get('total_size_mb', 0),
                    'file_types': summary.get('file_types', []),
                    'latest_file': summary.get('latest_file', None)
                }

        return jsonify(stats)
    except Exception as e:
        print(f"Error getting file stats: {e}")
        return jsonify({'error': 'File system error'}), 500

@app.post("/ckintelligence")
def ckintelligence_chat():
    """Handle chat messages for CK Intelligence."""
    # Check if user is logged in to either module
    if not (session.get('packing_logged_in') or session.get('raw_materials_logged_in')):
        return jsonify({'success': False, 'response': 'Please login to access CK Intelligence.'})

    try:
        data = request.get_json()
        if not data or 'message' not in data:
            return jsonify({'success': False, 'response': 'Invalid request format.'})

        user_message = data['message'].strip()
        explore_mode = data.get('explore_mode', False)

        # Get response from Google Gemini AI
        response = get_gemini_response(user_message, explore_mode)

        return jsonify({'success': True, 'response': response})

    except Exception as e:
        print(f"Error in ckintelligence_chat: {e}")
        return jsonify({'success': False, 'response': 'Sorry, I encountered an error processing your message. Please try again.'})


@app.get("/batch/<int:session_id>")
def batch_details(session_id: int):
    """View detailed information about a specific batch processing session."""
    # Check if user is logged in to either module
    if not (session.get('packing_logged_in') or session.get('raw_materials_logged_in')):
        flash('Please login to view batch details.', 'error')
        return redirect(url_for('index'))

    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # Get session information
    cursor.execute('''
        SELECT * FROM processing_sessions WHERE id = ?
    ''', (session_id,))
    session_data = cursor.fetchone()

    if not session_data:
        flash('Batch not found.', 'error')
        conn.close()
        return redirect(url_for('packing') if session.get('packing_logged_in') else url_for('raw_materials'))

    session_info = dict(session_data)

    # Get file information
    files = get_session_files(session_id)

    # Get processing results with order file information
    cursor.execute('''
        SELECT pr.*,
               uf.original_filename as order_filename
        FROM processing_results pr
        LEFT JOIN uploaded_files uf ON pr.order_file_id = uf.id
        WHERE pr.session_id = ?
        ORDER BY pr.fish_name ASC
    ''', (session_id,))
    results = [dict(row) for row in cursor.fetchall()]

    conn.close()

    # Debug: print results to see what we have
    print(f"Session {session_id} has {len(results)} processing results")
    if results:
        print(f"First result keys: {results[0].keys()}")

    return render_template('batch_details.html',
                         session_info=session_info,
                         files=files,
                         results=results)


@app.post("/process-batch")
def process_batch():
    stock_file = request.files.get("stock_file")
    import_stock_file = request.files.get("import_stock_file")
    extra_load_stock_file = request.files.get("extra_load_stock_file")
    order_files = request.files.getlist("order_files")
    if not stock_file or not order_files:
        flash("Please upload one Stock file and up to 32 Order files.", "error")
        return redirect(url_for("batch_index"))
    if len(order_files) > 32:
        flash("You can upload at most 32 order files.", "error")
        return redirect(url_for("batch_index"))

    try:
        # Load regular stock
        stock_ds = load_tabular(stock_file)

        # Combine all stock sources (regular, import, extra-load)
        combined_stock_rows = stock_ds.rows.copy()

        # Load import stock if provided and add to combined stock
        if import_stock_file and import_stock_file.filename:
            import_stock_ds = load_tabular(import_stock_file)
            combined_stock_rows.extend(import_stock_ds.rows)

        # Load extra-load stock if provided and add to combined stock
        if extra_load_stock_file and extra_load_stock_file.filename:
            extra_load_stock_ds = load_tabular(extra_load_stock_file)
            combined_stock_rows.extend(extra_load_stock_ds.rows)

        # Create a new dataset with combined rows if we have additional stock sources
        if len(combined_stock_rows) > len(stock_ds.rows):
            from dataclasses import dataclass
            @dataclass
            class CombinedDataset:
                rows: list
                name: str
                sheet_names: list

            stock_ds = CombinedDataset(
                rows=combined_stock_rows,
                name="combined_stock",
                sheet_names=["Combined Stock"]
            )
    except Exception as exc:  # noqa: BLE001
        flash(f"Failed to read stock file: {exc}", "error")
        return redirect(url_for("batch_index"))

    batch_token = uuid.uuid4().hex
    token_list: List[str] = []

    # Save batch session to database
    batch_session_id = None
    try:
        batch_session_id = save_processing_session(
            session_token=uuid.uuid4().hex,
            batch_token=batch_token,
            processing_type='batch'
        )
        
        # Save stock file for batch
        stock_file.seek(0)
        save_uploaded_file(batch_session_id, 'stock', stock_file.filename or 'stock.xlsx', stock_file)

        # Save import stock file if provided
        if import_stock_file and import_stock_file.filename:
            import_stock_file.seek(0)
            save_uploaded_file(batch_session_id, 'import_stock', import_stock_file.filename or 'import_stock.xlsx', import_stock_file)

        # Save extra-load stock file if provided
        if extra_load_stock_file and extra_load_stock_file.filename:
            extra_load_stock_file.seek(0)
            save_uploaded_file(batch_session_id, 'extra_load_stock', extra_load_stock_file.filename or 'extra_load_stock.xlsx', extra_load_stock_file)
    except Exception as e:
        print(f"Warning: Failed to save batch session: {e}")

    # Accumulate unique Not have fish across all orders in this batch
    fish_groups: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for of in order_files:
        try:
            order_ds = load_tabular(of)
            result_rows = compute_matches(stock_ds.rows, order_ds.rows)
            # Build summary like in single process
            def sum_required_kg(rows: List[Dict[str, Any]]) -> float:
                total = 0.0
                for r in rows:
                    try:
                        total += float(r.get("required_kg", 0) or 0)
                    except Exception:  # noqa: BLE001
                        pass
                return round(total, 3)

            # Compute ready kg from stock present for this order (sum over rows of stock_ctn * stock_kg_per_ctn)
            def sum_ready_kg(rows: List[Dict[str, Any]]) -> float:
                total = 0.0
                for r in rows:
                    try:
                        total += float(r.get("stock_carton", 0) or 0) * float(r.get("stock_kg_per_ctn", 0) or 0)
                    except Exception:  # noqa: BLE001
                        pass
                return round(total, 3)

            # Compute fulfillable kg using can_fulfill_carton * order_kg_per_ctn
            def sum_fulfillable_kg(rows: List[Dict[str, Any]]) -> float:
                total = 0.0
                for r in rows:
                    try:
                        can_fulfill = float(r.get("can_fulfill_carton", 0) or 0)
                        order_kg_per_ctn = float(r.get("order_kg_per_ctn", 0) or 0)
                        total += can_fulfill * order_kg_per_ctn
                    except Exception:  # noqa: BLE001
                        pass
                return round(total, 3)

            summary = {
                "total_items": int(len(result_rows)),
                "full": int(sum(1 for r in result_rows if r["status"] == "Full")),
                "not_full": int(sum(1 for r in result_rows if r["status"] == "Not Full")),
                "not_have": int(sum(1 for r in result_rows if r["status"] == "Not have")),
                "total_kg_all": sum_required_kg(result_rows),
                "total_kg_full": sum_required_kg([r for r in result_rows if r["status"] == "Full"]),
                "total_kg_not_full": sum_required_kg([r for r in result_rows if r["status"] == "Not Full"]),
                "total_kg_not_have": sum_required_kg([r for r in result_rows if r["status"] == "Not have"]),
                "ready_kg": sum_ready_kg(result_rows),
                "fulfillable_kg": sum_fulfillable_kg(result_rows),
            }

            # Build per-order Not have fish, aggregated by fish+pack with needed kg
            for row in result_rows:
                if row.get("status") == "Not have":
                    # Handle both underscore and space key formats for backward compatibility
                    fish_name = str(row.get("fish_name", row.get("fish name", "")))
                    pack_text = str(row.get("packed_size", row.get("packed size", "")))
                    needed_kg = float(row.get("required_kg", 0) or 0)
                    key = (fish_name, pack_text)
                    group = fish_groups.get(key)
                    if not group:
                        group = {
                            "fish_name": fish_name,
                            "packed_size": pack_text,
                            "total_needed_kg": 0.0,
                            "orders": [],
                        }
                        fish_groups[key] = group
                    group["total_needed_kg"] = round(group["total_needed_kg"] + needed_kg, 3)
                    # attach decision if already set
                    decision_key = f"{fish_name}|{pack_text}|{order_ds.name}"
                    decision = DECISION_STORE.get(decision_key)
                    group["orders"].append({
                        "order_name": order_ds.name,
                        "needed_kg": round(needed_kg, 3),
                        "decision": decision,
                    })

            # Save individual order processing to database
            try:
                if batch_session_id:
                    # Save order file and get its ID
                    of.seek(0)
                    order_file_stored_name = save_uploaded_file(batch_session_id, 'order', of.filename or 'order.xlsx', of)
                    
                    # Get the file ID for linking results
                    conn = sqlite3.connect(DATABASE_PATH)
                    cursor = conn.cursor()
                    cursor.execute('SELECT id FROM uploaded_files WHERE session_id = ? AND stored_filename = ?', 
                                 (batch_session_id, order_file_stored_name))
                    order_file_record = cursor.fetchone()
                    order_file_id = order_file_record[0] if order_file_record else None
                    conn.close()
                    
                    # Save processing results linked to this order file
                    save_processing_results(batch_session_id, result_rows, order_file_id)
                    # Persist any existing in-memory schedule for this token
                    # Note: New sessions won't have this yet; restored sessions will
                    scheduled_on = SCHEDULE_STORE.get(token)
                    if scheduled_on and order_file_id:
                        upsert_schedule(batch_session_id, order_file_id, scheduled_on)
            except Exception as e:
                print(f"Warning: Failed to save order processing: {e}")

            excel_bytes = rows_to_excel_bytes(result_rows)
            pdf_bytes = rows_to_pdf_bytes(result_rows)
            token = uuid.uuid4().hex
            order_basename = os.path.splitext(order_ds.name or "order")[0]
            # Store original order data for later editing
            original_order_data = []
            for row in order_ds.rows:
                normalized = try_map_row(row)
                weight_mc = normalized.get('weight_mc', '')
                
                # First try to parse order_kg_per_ctn from weight_mc
                try:
                    order_kg_per_ctn = float(weight_mc) if weight_mc not in (None, "") else 0.0
                except Exception:
                    order_kg_per_ctn = parse_kg_per_carton(weight_mc)
                
                # If we still don't have weight_mc but we parsed order_kg_per_ctn from packed size, use it
                if not weight_mc and order_kg_per_ctn == 0:
                    try:
                        order_kg_per_ctn = parse_kg_per_carton(normalized.get('packed size', ''))
                        if order_kg_per_ctn > 0:
                            weight_mc = str(order_kg_per_ctn)
                    except:
                        pass
                elif not weight_mc and order_kg_per_ctn > 0:
                    weight_mc = str(order_kg_per_ctn)
                
                order_row = {
                    'fish name': normalized.get('fish name', ''),
                    'packed size': normalized.get('packed size', ''),
                    'pack': normalized.get('pack', ''),
                    'total carton': to_int(normalized.get('total carton', 0)),
                    'weight_mc': weight_mc,
                    'order_kg_per_ctn': order_kg_per_ctn,
                    'remark': normalized.get('remark', '')
                }
                original_order_data.append(order_row)

            RESULT_STORE[token] = {
                "excel": excel_bytes,
                "pdf": pdf_bytes,
                "excel_name": f"{order_basename} Calculation.xlsx",
                "pdf_name": f"{order_basename} Calculation.pdf",
                "rows_json": io.BytesIO(str(result_rows).encode("utf-8")).getvalue(),
                "summary_json": io.BytesIO(str(summary).encode("utf-8")).getvalue(),
                "stock_name": stock_ds.name.encode("utf-8"),
                "order_name": order_ds.name.encode("utf-8"),
                "original_order_json": io.BytesIO(str(original_order_data).encode("utf-8")).getvalue(),
            }
            token_list.append(token)
        except Exception as exc:  # noqa: BLE001
            flash(f"Failed to process order file {getattr(of,'filename','unknown')}: {exc}", "error")

    BATCH_STORE[batch_token] = token_list

    # Load persisted schedules from DB and decorate items
    try:
        schedule_map = get_schedules_for_batch(batch_token)
    except Exception:
        schedule_map = {}

    # Build summary items
    items = []
    for t in token_list:
        entry = RESULT_STORE.get(t, {})
        try:
            # Decode stored summary/metadata
            order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
            summary = eval((entry.get("summary_json") or b"{}"))  # noqa: S307
            # If ready_kg missing (older entry), compute from rows
            if "ready_kg" not in summary:
                try:
                    rows_tmp = eval((entry.get("rows_json") or b"[]"))  # noqa: S307
                    ready = 0.0
                    for rr in rows_tmp:
                        ready += float(rr.get("stock_carton", 0) or 0) * float(rr.get("stock_kg_per_ctn", 0) or 0)
                    summary["ready_kg"] = round(ready, 3)
                except Exception:
                    summary["ready_kg"] = 0.0

            items.append({
                "token": t,
                "order_name": order_name,
                "summary": summary,
                "scheduled_on": schedule_map.get(order_name) or SCHEDULE_STORE.get(t),
            })
        except Exception:
            continue

    # Build chart data
    labels = [it["order_name"] for it in items]
    full_counts = [it["summary"].get("full", 0) for it in items]
    not_full_counts = [it["summary"].get("not_full", 0) for it in items]
    not_have_counts = [it["summary"].get("not_have", 0) for it in items]

    kg_full = sum(it["summary"].get("total_kg_full", 0) for it in items)
    kg_not_full = sum(it["summary"].get("total_kg_not_full", 0) for it in items)
    kg_not_have = sum(it["summary"].get("total_kg_not_have", 0) for it in items)

    events = []
    for it in items:
        if it.get("scheduled_on"):
            events.append({
                "title": it["order_name"],
                "start": it["scheduled_on"],
                "url": url_for('view_result', token=it["token"]),
                "extendedProps": {"token": it["token"]},
            })

    # Turn fish groups dict into a list for template
    fish_groups_list = [
        {
            "fish_name": k[0],
            "packed_size": k[1],
            "total_needed_kg": v["total_needed_kg"],
            "orders": v["orders"],
        }
        for k, v in fish_groups.items()
    ]

    fish_total_kg = round(sum(g["total_needed_kg"] for g in fish_groups_list), 3)

    # Recommendations: sort orders by number of Full items desc; include ready_kg
    recommendations = sorted(
        (
            {
                "order_name": it["order_name"],
                "full": it["summary"].get("full", 0),
                "full_kg": it["summary"].get("total_kg_full", 0),
                "token": it["token"],
                "scheduled": bool(it.get("scheduled_on")),
            }
            for it in items
        ),
        key=lambda x: (x["full"], x["full_kg"]),
        reverse=True,
    )

    # Persist aggregated batch summary numbers for accurate history display
    try:
        if batch_session_id:
            agg_total_items = sum(it["summary"].get("total_items", 0) for it in items)
            agg_full = sum(it["summary"].get("full", 0) for it in items)
            agg_nf = sum(it["summary"].get("not_full", 0) for it in items)
            agg_nh = sum(it["summary"].get("not_have", 0) for it in items)
            agg_kg_all = sum(it["summary"].get("total_kg_all", 0) for it in items)
            agg_kg_full = sum(it["summary"].get("total_kg_full", 0) for it in items)
            agg_kg_nf = sum(it["summary"].get("total_kg_not_full", 0) for it in items)
            agg_kg_nh = sum(it["summary"].get("total_kg_not_have", 0) for it in items)

            conn = sqlite3.connect(DATABASE_PATH)
            cursor = conn.cursor()
            cursor.execute(
                '''UPDATE processing_sessions
                   SET total_items=?, full_items=?, not_full_items=?, not_have_items=?,
                       total_kg=?, full_kg=?, not_full_kg=?, not_have_kg=?
                 WHERE id=?''',
                (
                    int(agg_total_items), int(agg_full), int(agg_nf), int(agg_nh),
                    float(agg_kg_all), float(agg_kg_full), float(agg_kg_nf), float(agg_kg_nh),
                    batch_session_id,
                )
            )
            conn.commit()
            conn.close()
    except Exception:
        pass

    return redirect(url_for('view_batch', batch_token=batch_token))


@app.get("/get-comparison-history/<batch_token>")
def get_comparison_history(batch_token: str):
    """Get file comparison history for a batch."""
    try:
        history = get_file_comparison_history(batch_token)
        return {"success": True, "history": history}
    except Exception as e:
        return {"success": False, "error": str(e)}, 500

@app.post("/clear-comparison-history/<batch_token>")
def clear_comparison_history(batch_token: str):
    """Clear comparison history for a batch (for testing/debugging)."""
    try:
        session_id = get_session_id_by_batch_token(batch_token)
        if session_id:
            conn = sqlite3.connect(DATABASE_PATH)
            cursor = conn.cursor()
            cursor.execute('DELETE FROM file_comparison_history WHERE session_id = ?', (session_id,))
            conn.commit()
            conn.close()
            return {"success": True, "message": "History cleared"}
        else:
            return {"success": False, "error": "Session not found"}, 404
    except Exception as e:
        return {"success": False, "error": str(e)}, 500

@app.get("/batch/<batch_token>")
def view_batch(batch_token: str):
    tokens = BATCH_STORE.get(batch_token)
    if not tokens:
        flash("Unknown or expired batch.", "error")
        return redirect(url_for("batch_index"))

    # Get batch session ID and load all stock data
    batch_session_id = get_session_id_by_batch_token(batch_token)
    import_stock_lookup = {}

    if batch_session_id:
        # Load all stock files (import and extra-load) and create lookup
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Load both import stock and extra-load stock files
        cursor.execute('''
            SELECT stored_filename, file_type FROM uploaded_files
            WHERE session_id = ? AND file_type IN ('import_stock', 'extra_load_stock')
        ''', (batch_session_id,))

        for row in cursor.fetchall():
            try:
                stock_path = os.path.join(UPLOAD_FOLDER, row['stored_filename'])
                if os.path.exists(stock_path):
                    with open(stock_path, 'rb') as f:
                        file_like = io.BytesIO(f.read())
                        file_like.filename = stock_path
                        stock_ds = load_tabular(file_like)

                        # Create lookup by (fish_name, packed_size)
                        for stock_row in stock_ds.rows:
                            try:
                                # Handle both dictionary and other data structures
                                if isinstance(stock_row, dict):
                                    fish_name = str(stock_row.get("fish_name", stock_row.get("fish name", "")))
                                    pack_size = str(stock_row.get("packed_size", stock_row.get("packed size", "")))
                                    qty = to_int(stock_row.get("total carton", 0))
                                else:
                                    # Fallback for other data structures
                                    continue

                                if fish_name and pack_size:
                                    key = (fish_name, pack_size)
                                    if key in import_stock_lookup:
                                        import_stock_lookup[key] += qty
                                    else:
                                        import_stock_lookup[key] = qty
                            except (AttributeError, KeyError, TypeError) as e:
                                # Skip problematic rows
                                print(f"Warning: Skipping stock row due to error: {e}")
                                continue
            except Exception as e:
                print(f"Warning: Failed to load stock file {row['stored_filename']}: {e}")

        conn.close()

    items = []
    fish_groups: Dict[Tuple[str, str], Dict[str, Any]] = {}
    stock_name = ""
    for t in tokens:
        entry = RESULT_STORE.get(t, {})
        try:
            order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
            summary = eval((entry.get("summary_json") or b"{}"))  # noqa: S307
            stock_name = (entry.get("stock_name") or b"").decode("utf-8", errors="ignore") or stock_name
            # aggregate Not have by fish+pack with needed kg per order
            try:
                rows = eval((entry.get("rows_json") or b"[]"))  # noqa: S307
                for row in rows:
                    if row.get("status") == "Not have":
                        # Handle both underscore and space key formats for backward compatibility
                        fish_name = str(row.get("fish_name", row.get("fish name", "")))
                        pack_text = str(row.get("packed_size", row.get("packed size", "")))
                        needed_kg = float(row.get("required_kg", 0) or 0)
                        key = (fish_name, pack_text)

                        # Skip items that are available in import stock
                        if key in import_stock_lookup and import_stock_lookup[key] > 0:
                            continue

                        group = fish_groups.get(key)
                        if not group:
                            group = {
                                "fish_name": fish_name,
                                "packed_size": pack_text,
                                "total_needed_kg": 0.0,
                                "orders": [],
                            }
                            fish_groups[key] = group
                        group["total_needed_kg"] = round(group["total_needed_kg"] + needed_kg, 3)
                        decision_key = f"{fish_name}|{pack_text}|{order_name}"
                        decision = DECISION_STORE.get(decision_key)
                        group["orders"].append({
                            "order_name": order_name,
                            "needed_kg": round(needed_kg, 3),
                            "decision": decision,
                        })
            except Exception:
                pass

            items.append({
                "token": t,
                "order_name": order_name,
                "summary": summary,
                "scheduled_on": get_schedules_for_batch(batch_token).get(order_name) or SCHEDULE_STORE.get(t),
            })
        except Exception:
            continue

    labels = [it["order_name"] for it in items]
    full_counts = [it["summary"].get("full", 0) for it in items]
    not_full_counts = [it["summary"].get("not_full", 0) for it in items]
    not_have_counts = [it["summary"].get("not_have", 0) for it in items]
    kg_full = sum(it["summary"].get("total_kg_full", 0) for it in items)
    kg_not_full = sum(it["summary"].get("total_kg_not_full", 0) for it in items)
    kg_not_have = sum(it["summary"].get("total_kg_not_have", 0) for it in items)
    events = []
    for it in items:
        if it.get("scheduled_on"):
            events.append({
                "title": it["order_name"],
                "start": it["scheduled_on"],
                "url": url_for('view_result', token=it["token"]),
                "extendedProps": {"token": it["token"]},
            })

    fish_groups_list = [
        {
            "fish_name": k[0],
            "packed_size": k[1],
            "total_needed_kg": v["total_needed_kg"],
            "orders": v["orders"],
        }
        for k, v in fish_groups.items()
    ]

    fish_total_kg = round(sum(g["total_needed_kg"] for g in fish_groups_list), 3)

    recommendations = sorted(
        (
            {
                "order_name": it["order_name"],
                "full": it["summary"].get("full", 0),
                "full_kg": it["summary"].get("total_kg_full", 0),
                "token": it["token"],
                "scheduled": bool(it.get("scheduled_on")),
            }
            for it in items
        ),
        key=lambda x: (x["full"], x["full_kg"]),
        reverse=True,
    )

    finished_orders = get_finished_orders_for_batch(batch_token)

    return render_template(
        "summary.html",
        batch_token=batch_token,
        stock_name=stock_name,
        items=items,
        fish_groups=fish_groups_list,
        fish_total_kg=fish_total_kg,
        recommendations=recommendations,
        chart_labels=json.dumps(labels),
        chart_full=json.dumps(full_counts),
        chart_not_full=json.dumps(not_full_counts),
        chart_not_have=json.dumps(not_have_counts),
        doughnut_data=json.dumps([kg_full, kg_not_full, kg_not_have]),
        calendar_events=json.dumps(events),
        finished_orders=finished_orders,
    )


@app.post("/set-decision")
def set_decision():
    fish_name = request.form.get('fish_name') or ''
    packed_size = request.form.get('packed_size') or ''
    order_name = request.form.get('order_name') or ''
    decision = request.form.get('decision') or ''
    batch_token = request.form.get('batch_token') or ''
    redirect_to = request.form.get('redirect_to') or ''
    key = f"{fish_name}|{packed_size}|{order_name}"
    if decision:
        DECISION_STORE[key] = decision
        # Persist to DB (best-effort): attach to most recent batch session
        try:
            conn = sqlite3.connect(DATABASE_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM processing_sessions WHERE processing_type='batch' ORDER BY created_at DESC LIMIT 1")
            row = cursor.fetchone()
            if row:
                upsert_fish_decision(row['id'], fish_name, packed_size, order_name, decision)
            conn.close()
        except Exception:
            pass
    else:
        DECISION_STORE.pop(key, None)
    if batch_token:
        if redirect_to == 'fish_buy':
            return redirect(url_for('fish_buy', batch_token=batch_token))
        # default: bring user back to summary with fish tab active (handled by JS init)
        return redirect(url_for('view_batch', batch_token=batch_token))
    return redirect(url_for('batch_index'))


def build_fish_groups_from_batch(batch_token: str) -> List[Dict[str, Any]]:
    tokens = BATCH_STORE.get(batch_token, [])

    # Get batch session ID and load all stock data
    batch_session_id = get_session_id_by_batch_token(batch_token)
    import_stock_lookup = {}

    if batch_session_id:
        # Load all stock files (import and extra-load) and create lookup
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Load both import stock and extra-load stock files
        cursor.execute('''
            SELECT stored_filename, file_type FROM uploaded_files
            WHERE session_id = ? AND file_type IN ('import_stock', 'extra_load_stock')
        ''', (batch_session_id,))

        for row in cursor.fetchall():
            try:
                stock_path = os.path.join(UPLOAD_FOLDER, row['stored_filename'])
                if os.path.exists(stock_path):
                    with open(stock_path, 'rb') as f:
                        file_like = io.BytesIO(f.read())
                        file_like.filename = stock_path
                        stock_ds = load_tabular(file_like)

                        # Create lookup by (fish_name, packed_size)
                        for stock_row in stock_ds.rows:
                            try:
                                # Handle both dictionary and other data structures
                                if isinstance(stock_row, dict):
                                    fish_name = str(stock_row.get("fish_name", stock_row.get("fish name", "")))
                                    pack_size = str(stock_row.get("packed_size", stock_row.get("packed size", "")))
                                    qty = to_int(stock_row.get("total carton", 0))
                                else:
                                    # Fallback for other data structures
                                    continue

                                if fish_name and pack_size:
                                    key = (fish_name, pack_size)
                                    if key in import_stock_lookup:
                                        import_stock_lookup[key] += qty
                                    else:
                                        import_stock_lookup[key] = qty
                            except (AttributeError, KeyError, TypeError) as e:
                                # Skip problematic rows
                                print(f"Warning: Skipping stock row due to error: {e}")
                                continue
            except Exception as e:
                print(f"Warning: Failed to load stock file {row['stored_filename']}: {e}")

        conn.close()

    fish_groups: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for t in tokens:
        entry = RESULT_STORE.get(t, {})
        try:
            order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
            rows = eval((entry.get("rows_json") or b"[]"))  # noqa: S307
            for row in rows:
                if row.get("status") == "Not have":
                    # Handle both underscore and space key formats for backward compatibility
                    fish_name = str(row.get("fish_name", row.get("fish name", "")))
                    pack_text = str(row.get("packed_size", row.get("packed size", "")))
                    needed_kg = float(row.get("required_kg", 0) or 0)
                    key = (fish_name, pack_text)

                    # Skip items that are available in import stock
                    if key in import_stock_lookup and import_stock_lookup[key] > 0:
                        continue

                    group = fish_groups.get(key)
                    if not group:
                        group = {
                            "fish_name": fish_name,
                            "packed_size": pack_text,
                            "total_needed_kg": 0.0,
                            "orders": [],
                        }
                        fish_groups[key] = group
                    group["total_needed_kg"] = round(group["total_needed_kg"] + needed_kg, 3)
                    decision_key = f"{fish_name}|{pack_text}|{order_name}"
                    decision = DECISION_STORE.get(decision_key)
                    group["orders"].append({
                        "order_name": order_name,
                        "needed_kg": round(needed_kg, 3),
                        "decision": decision,
                    })
        except Exception:
            continue
    return [
        {
            "fish_name": k[0],
            "packed_size": k[1],
            "total_needed_kg": v["total_needed_kg"],
            "orders": v["orders"],
        }
        for k, v in fish_groups.items()
    ]


@app.get("/fish-buy/<batch_token>")
def fish_buy(batch_token: str):
    # reuse batch view data
    tokens = BATCH_STORE.get(batch_token)
    if not tokens:
        flash("Unknown or expired batch.", "error")
        return redirect(url_for("batch_index"))

    # stock name from first token
    stock_name = ""
    for t in tokens:
        entry = RESULT_STORE.get(t, {})
        stock_name = (entry.get("stock_name") or b"").decode("utf-8", errors="ignore")
        if stock_name:
            break

    items = []
    for t in tokens:
        entry = RESULT_STORE.get(t, {})
        try:
            order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
            summary = eval((entry.get("summary_json") or b"{}"))  # noqa: S307
            items.append({"token": t, "order_name": order_name, "summary": summary})
        except Exception:
            continue

    fish_groups = build_fish_groups_from_batch(batch_token)

    # charts data (optional on this page)
    labels = [it["order_name"] for it in items]
    full_counts = [it["summary"].get("full", 0) for it in items]
    not_full_counts = [it["summary"].get("not_full", 0) for it in items]
    not_have_counts = [it["summary"].get("not_have", 0) for it in items]
    kg_full = sum(it["summary"].get("total_kg_full", 0) for it in items)
    kg_not_full = sum(it["summary"].get("total_kg_not_full", 0) for it in items)
    kg_not_have = sum(it["summary"].get("total_kg_not_have", 0) for it in items)

    return render_template(
        "fish_buy.html",
        batch_token=batch_token,
        stock_name=stock_name,
        items=items,
        fish_groups=fish_groups,
        chart_labels=json.dumps(labels),
        chart_full=json.dumps(full_counts),
        chart_not_full=json.dumps(not_full_counts),
        chart_not_have=json.dumps(not_have_counts),
        doughnut_data=json.dumps([kg_full, kg_not_full, kg_not_have]),
    )


def fish_groups_to_excel_bytes(fish_groups: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fish Summary"
    ws.append(["Fish Name", "Packed Size", "Total Needed KG", "Orders Count"])
    for g in fish_groups:
        ws.append([g.get("fish_name", ""), g.get("packed_size", ""), g.get("total_needed_kg", 0), len(g.get("orders", []))])
    ws.freeze_panes = "A2"

    # Decisions sheet: only orders with a decision
    ws2 = wb.create_sheet("Decisions")
    ws2.append(["Fish Name", "Packed Size", "Order File", "Needed KG", "Decision"])
    for g in fish_groups:
        for o in g.get("orders", []):
            if o.get("decision"):
                ws2.append([g.get("fish_name", ""), g.get("packed_size", ""), o.get("order_name", ""), o.get("needed_kg", 0), o.get("decision")])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@app.get("/download/fish-excel")
def download_fish_excel():
    batch_token = request.args.get('batch', type=str)
    if not batch_token or batch_token not in BATCH_STORE:
        flash("Unknown batch.", "error")
        return redirect(url_for('batch_index'))
    fish_groups = build_fish_groups_from_batch(batch_token)
    raw = fish_groups_to_excel_bytes(fish_groups)
    return send_file(
        io.BytesIO(raw),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"Fish Decisions - {batch_token}.xlsx",
    )


@app.post("/schedule-order")
def schedule_order():
    token = request.form.get('token')
    date = request.form.get('date')
    batch_token = request.form.get('batch_token')
    if token and date:
        SCHEDULE_STORE[token] = date
        # Persist to DB if this token belongs to a batch restored session
        # Find the order name from RESULT_STORE
        entry = RESULT_STORE.get(token)
        if entry and batch_token:
            order_name = (entry.get('order_name') or b'').decode('utf-8', errors='ignore')
            stock_name = (entry.get('stock_name') or b'').decode('utf-8', errors='ignore')
            # Locate a session that matches this stock file and batch (best-effort)
            try:
                conn = sqlite3.connect(DATABASE_PATH)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                # Find exact session by batch_token
                cursor.execute("SELECT id FROM processing_sessions WHERE batch_token = ? LIMIT 1", (batch_token,))
                row = cursor.fetchone()
                if row:
                    session_id = row['id']
                    # Find order file id by original filename
                    cursor.execute("SELECT id FROM uploaded_files WHERE session_id=? AND file_type='order' AND original_filename=? ORDER BY id DESC LIMIT 1", (session_id, order_name))
                    of = cursor.fetchone()
                    if of:
                        upsert_schedule(session_id, of['id'], date)
                conn.close()
            except Exception:
                pass
    if batch_token:
        return redirect(url_for('view_batch', batch_token=batch_token))
    return redirect(url_for('batch_index'))


@app.post("/unschedule-order")
def unschedule_order():
    token = request.form.get('token')
    batch_token = request.form.get('batch_token')
    if token:
        SCHEDULE_STORE.pop(token, None)
        # Remove from DB as well (best-effort)
        entry = RESULT_STORE.get(token)
        if entry and batch_token:
            order_name = (entry.get('order_name') or b'').decode('utf-8', errors='ignore')
            try:
                conn = sqlite3.connect(DATABASE_PATH)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM processing_sessions WHERE batch_token = ? LIMIT 1", (batch_token,))
                row = cursor.fetchone()
                if row:
                    session_id = row['id']
                    cursor.execute("SELECT id FROM uploaded_files WHERE session_id=? AND file_type='order' AND original_filename=? ORDER BY id DESC LIMIT 1", (session_id, order_name))
                    of = cursor.fetchone()
                    if of:
                        delete_schedule(session_id, of['id'])
                conn.close()
            except Exception:
                pass
    # For fetch usage, return a simple OK
    return "OK"


@app.post('/remove-order')
def remove_order():
    token = request.form.get('token')
    batch_token = request.form.get('batch_token')
    if not token or not batch_token:
        flash('Invalid remove request.', 'error')
        return redirect(url_for('batch_index'))

    # Capture entry (order name) before mutating in-memory stores
    entry_snapshot = RESULT_STORE.get(token)
    order_name_snapshot = (entry_snapshot.get('order_name') or b'').decode('utf-8', errors='ignore') if entry_snapshot else None

    # Remove from in-memory batch store
    tokens = BATCH_STORE.get(batch_token, [])
    if token in tokens:
        tokens.remove(token)
        BATCH_STORE[batch_token] = tokens
    # Also drop in-memory result for this token
    RESULT_STORE.pop(token, None)

    # Best-effort: remove the order file record and its results from DB for this batch
    try:
        order_name = order_name_snapshot
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM processing_sessions WHERE batch_token = ? LIMIT 1', (batch_token,))
        row = cursor.fetchone()
        if row and order_name:
            session_id = row['id']
            cursor.execute("SELECT id FROM uploaded_files WHERE session_id=? AND file_type='order' AND original_filename=?", (session_id, order_name))
            of = cursor.fetchone()
            if of:
                order_file_id = of['id']
                cursor.execute('DELETE FROM processing_results WHERE session_id=? AND order_file_id=?', (session_id, order_file_id))
                cursor.execute('DELETE FROM scheduled_orders WHERE session_id=? AND order_file_id=?', (session_id, order_file_id))
                cursor.execute('DELETE FROM uploaded_files WHERE id=?', (order_file_id,))
        conn.commit()
        conn.close()
    except Exception:
        pass

    flash('Order removed.', 'success')
    return redirect(url_for('view_batch', batch_token=batch_token))


@app.get("/result/<token>")
def view_result(token: str):
    entry = RESULT_STORE.get(token)
    if not entry:
        flash("Unknown or expired result token.", "error")
        return redirect(url_for("index"))
    try:
        rows = eval(entry.get("rows_json", b"[]"))  # noqa: S307
        summary = eval(entry.get("summary_json", b"{}"))  # noqa: S307
        stock_name = (entry.get("stock_name") or b"").decode("utf-8", errors="ignore")
        order_name = (entry.get("order_name") or b"").decode("utf-8", errors="ignore")
    except Exception:
        flash("Failed to load stored result.", "error")
        return redirect(url_for("index"))

    return render_template(
        "result.html",
        summary=summary,
        records=rows,
        stock_name=stock_name,
        order_name=order_name,
        download_token=token,
    )


@app.get("/download/excel")
def download_excel():
    token = request.args.get("token", type=str)
    raw = RESULT_STORE.get(token or "", {}).get("excel") if token else None
    if not raw:
        flash("No processed result to download yet.", "error")
        return redirect(url_for("index"))
    name = RESULT_STORE.get(token or "", {}).get("excel_name", "order_stock_result.xlsx")
    return send_file(
        io.BytesIO(raw),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=name,
    )


@app.get("/download/pdf")
def download_pdf():
    token = request.args.get("token", type=str)
    raw = RESULT_STORE.get(token or "", {}).get("pdf") if token else None
    if not raw:
        flash("No processed result to download yet.", "error")
        return redirect(url_for("index"))
    name = RESULT_STORE.get(token or "", {}).get("pdf_name", "order_stock_result.pdf")
    return send_file(
        io.BytesIO(raw),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=name,
    )


@app.get("/session/<int:session_id>")
def view_session(session_id: int):
    """View a saved processing session."""
    session = get_session_by_id(session_id)
    if not session:
        flash("Session not found.", "error")
        return redirect(url_for("index"))
    
    # For batch sessions, restore to memory and redirect to batch summary
    if session.get('processing_type') == 'batch':
        batch_token = restore_batch_session_to_memory(session)
        if batch_token:
            return redirect(url_for('view_batch', batch_token=batch_token))
        else:
            flash("Failed to restore batch session.", "error")
            return redirect(url_for("index"))
    
    # Handle single sessions
    results = session.get('results', [])
    records = []
    for r in results:
        record = {
            'fish name': r.get('fish_name', ''),
            'packed size': r.get('packed_size', ''),
            'order_carton': r.get('order_carton', 0),
            'stock_carton': r.get('stock_carton', 0),
            'order_kg_per_ctn': r.get('order_kg_per_ctn', 0),
            'stock_kg_per_ctn': r.get('stock_kg_per_ctn', 0),
            'balance_stock_carton': r.get('balance_stock_carton', 0),
            'mc_to_give': r.get('mc_to_give', 0),
            'can_fulfill_carton': r.get('can_fulfill_carton', 0),
            'shortfall': r.get('shortfall', 0),
            'status': r.get('status', ''),
            'required_kg': r.get('required_kg', 0)
        }
        records.append(record)
    
    # Create summary from session data
    summary = {
        'total_items': session.get('total_items', 0),
        'full': session.get('full_items', 0),
        'not_full': session.get('not_full_items', 0),
        'not_have': session.get('not_have_items', 0),
        'total_kg_all': session.get('total_kg', 0),
        'total_kg_full': session.get('full_kg', 0),
        'total_kg_not_full': session.get('not_full_kg', 0),
        'total_kg_not_have': session.get('not_have_kg', 0)
    }
    
    return render_template(
        "result.html",
        summary=summary,
        records=records,
        stock_name=session.get('stock_filename', 'Unknown Stock File'),
        order_name=session.get('order_filenames', 'Unknown Order File'),
        download_token=None,  # No download available for historical data
        is_historical=True
    )


@app.post("/compare-order-files")
def compare_order_files():
    """Compare an original order file with an updated version and return differences."""
    try:
        original_token = request.form.get('original_token')
        batch_token = request.form.get('batch_token')
        updated_file = request.files.get('updated_file')
        
        if not original_token or not updated_file:
            return {"success": False, "error": "Missing required parameters"}, 400
        
        # Get original file data from RESULT_STORE
        original_entry = RESULT_STORE.get(original_token)
        if not original_entry:
            return {"success": False, "error": "Original file not found"}, 404
        
        # Get original filename for validation
        try:
            original_filename = (original_entry.get("order_name") or b"").decode("utf-8", errors="ignore")
        except:
            original_filename = ""
        
        # Validate filename matches (base name without extension)
        if original_filename and updated_file.filename:
            import os
            original_base = os.path.splitext(original_filename)[0].lower()
            updated_base = os.path.splitext(updated_file.filename)[0].lower()
            
            if original_base != updated_base:
                return {
                    "success": False, 
                    "error": f"File name mismatch. Expected: {original_filename}, Got: {updated_file.filename}"
                }, 400
        
        # Parse original file data - use stored original order data if available
        try:
            if "original_order_json" in original_entry:
                # Use stored original order data
                original_rows = eval(original_entry.get("original_order_json", b"[]"))
            else:
                # Fallback: reconstruct from processed results
                processed_rows = eval(original_entry.get("rows_json", b"[]"))
                original_rows = []
                for row in processed_rows:
                    order_row = {
                        'fish name': row.get('fish name', ''),
                        'packed size': row.get('packed size', ''),
                        'total carton': row.get('order_carton', 0),
                        'weight_mc': '',
                        'order_kg_per_ctn': row.get('order_kg_per_ctn', 0)
                    }
                    original_rows.append(order_row)
        except:
            return {"success": False, "error": "Could not parse original file data"}, 500
        
        # Load and parse updated file
        try:
            updated_ds = load_tabular(updated_file)
            updated_rows = []
            for row in updated_ds.rows:
                normalized = try_map_row(row)
                weight_mc = normalized.get('weight_mc', '')
                
                # Parse order_kg_per_ctn from weight_mc
                try:
                    order_kg_per_ctn = float(weight_mc) if weight_mc not in (None, "") else 0.0
                except Exception:
                    order_kg_per_ctn = parse_kg_per_carton(weight_mc)
                
                # If we still don't have weight_mc but we parsed order_kg_per_ctn from packed size, use it
                if not weight_mc and order_kg_per_ctn == 0:
                    try:
                        order_kg_per_ctn = parse_kg_per_carton(normalized.get('packed size', ''))
                        if order_kg_per_ctn > 0:
                            weight_mc = str(order_kg_per_ctn)
                    except:
                        pass
                elif not weight_mc and order_kg_per_ctn > 0:
                    weight_mc = str(order_kg_per_ctn)
                
                updated_rows.append({
                    'fish name': normalized.get('fish name', ''),
                    'packed size': normalized.get('packed size', ''),
                    'pack': normalized.get('pack', ''),
                    'total carton': to_int(normalized.get('total carton', 0)),
                    'weight_mc': weight_mc,
                    'order_kg_per_ctn': order_kg_per_ctn,
                    'remark': normalized.get('remark', '')
                })
        except Exception as e:
            return {"success": False, "error": f"Could not parse updated file: {str(e)}"}, 400
        
        # Perform comparison
        comparison_result = compare_order_file_data(original_rows, updated_rows)
        
        # Don't save comparison history here - only save when changes are actually applied
        # This prevents duplicate entries in the history table
        
        return {
            "success": True,
            "comparison": comparison_result
        }
        
    except Exception as e:
        return {"success": False, "error": str(e)}, 500


def compare_order_file_data(original_rows, updated_rows):
    """Compare two sets of order data and return detailed differences."""
    # Create lookup dictionaries for comparison
    original_lookup = {}
    updated_lookup = {}
    
    # Build lookup for original data
    for row in original_rows:
        key = (
            canonicalize_product(row.get('fish name', '')),
            canonicalize_pack(row.get('packed size', ''))
        )
        original_lookup[key] = {
            'fish_name': row.get('fish name', ''),
            'packed_size': row.get('packed size', ''),
            'pack': row.get('pack', ''),
            'quantity': row.get('total carton', 0) if 'total carton' in row else row.get('order_carton', 0),
            'weight_mc': row.get('weight_mc', ''),
            'order_kg_per_ctn': row.get('order_kg_per_ctn', 0),
            'remark': row.get('remark', ''),
            'raw_data': row
        }
    
    # Build lookup for updated data
    for row in updated_rows:
        key = (
            canonicalize_product(row.get('fish name', '')),
            canonicalize_pack(row.get('packed size', ''))
        )
        updated_lookup[key] = {
            'fish_name': row.get('fish name', ''),
            'packed_size': row.get('packed size', ''),
            'pack': row.get('pack', ''),
            'quantity': row.get('total carton', 0),
            'weight_mc': row.get('weight_mc', ''),
            'order_kg_per_ctn': row.get('order_kg_per_ctn', 0),
            'remark': row.get('remark', ''),
            'raw_data': row
        }
    
    # Find changes
    changes = []
    all_keys = set(original_lookup.keys()) | set(updated_lookup.keys())
    
    added_count = 0
    modified_count = 0
    deleted_count = 0
    unchanged_count = 0
    
    for key in all_keys:
        original_item = original_lookup.get(key)
        updated_item = updated_lookup.get(key)
        
        if original_item and updated_item:
            # Item exists in both - check for modifications
            quantity_changed = original_item['quantity'] != updated_item['quantity']
            weight_changed = str(original_item['weight_mc']) != str(updated_item['weight_mc'])
            pack_changed = str(original_item['pack']) != str(updated_item['pack'])
            remark_changed = str(original_item['remark']) != str(updated_item['remark'])
            
            if quantity_changed or weight_changed or pack_changed or remark_changed:
                change_details = []
                if quantity_changed:
                    change_details.append(f"Quantity: {original_item['quantity']} → {updated_item['quantity']}")
                if weight_changed:
                    change_details.append(f"Weight MC: {original_item['weight_mc']} → {updated_item['weight_mc']}")
                if pack_changed:
                    change_details.append(f"Pack: {original_item['pack']} → {updated_item['pack']}")
                if remark_changed:
                    change_details.append(f"Remark: {original_item['remark']} → {updated_item['remark']}")
                
                changes.append({
                    'status': 'modified',
                    'fish_name': updated_item['fish_name'],
                    'packed_size': updated_item['packed_size'],
                    'old_quantity': original_item['quantity'],
                    'new_quantity': updated_item['quantity'],
                    'old_weight_mc': original_item['weight_mc'],
                    'new_weight_mc': updated_item['weight_mc'],
                    'old_pack': original_item['pack'],
                    'new_pack': updated_item['pack'],
                    'old_remark': original_item['remark'],
                    'new_remark': updated_item['remark'],
                    'old_order_kg_per_ctn': original_item['order_kg_per_ctn'],
                    'new_order_kg_per_ctn': updated_item['order_kg_per_ctn'],
                    'changes': "; ".join(change_details)
                })
                modified_count += 1
            else:
                changes.append({
                    'status': 'unchanged',
                    'fish_name': updated_item['fish_name'],
                    'packed_size': updated_item['packed_size'],
                    'old_quantity': original_item['quantity'],
                    'new_quantity': updated_item['quantity'],
                    'old_weight_mc': original_item['weight_mc'],
                    'new_weight_mc': updated_item['weight_mc'],
                    'old_pack': original_item['pack'],
                    'new_pack': updated_item['pack'],
                    'old_remark': original_item['remark'],
                    'new_remark': updated_item['remark'],
                    'old_order_kg_per_ctn': original_item['order_kg_per_ctn'],
                    'new_order_kg_per_ctn': updated_item['order_kg_per_ctn'],
                    'changes': 'No changes'
                })
                unchanged_count += 1
        elif updated_item and not original_item:
            # New item added
            changes.append({
                'status': 'added',
                'fish_name': updated_item['fish_name'],
                'packed_size': updated_item['packed_size'],
                'old_quantity': None,
                'new_quantity': updated_item['quantity'],
                'old_weight_mc': None,
                'new_weight_mc': updated_item['weight_mc'],
                'old_pack': None,
                'new_pack': updated_item['pack'],
                'old_remark': None,
                'new_remark': updated_item['remark'],
                'old_order_kg_per_ctn': None,
                'new_order_kg_per_ctn': updated_item['order_kg_per_ctn'],
                'changes': 'New item added'
            })
            added_count += 1
        elif original_item and not updated_item:
            # Item deleted
            changes.append({
                'status': 'deleted',
                'fish_name': original_item['fish_name'],
                'packed_size': original_item['packed_size'],
                'old_quantity': original_item['quantity'],
                'new_quantity': None,
                'old_weight_mc': original_item['weight_mc'],
                'new_weight_mc': None,
                'old_pack': original_item['pack'],
                'new_pack': None,
                'old_remark': original_item['remark'],
                'new_remark': None,
                'old_order_kg_per_ctn': original_item['order_kg_per_ctn'],
                'new_order_kg_per_ctn': None,
                'changes': 'Item removed'
            })
            deleted_count += 1
    
    # Sort changes by status and fish name
    changes.sort(key=lambda x: (
        {'added': 0, 'modified': 1, 'deleted': 2, 'unchanged': 3}[x['status']],
        x['fish_name'].lower()
    ))
    
    return {
        'summary': {
            'added': added_count,
            'modified': modified_count,
            'deleted': deleted_count,
            'unchanged': unchanged_count
        },
        'changes': changes
    }


@app.post("/apply-order-changes")
def apply_order_changes():
    """Apply changes from comparison back to the original order file and update results."""
    try:
        original_token = request.form.get('original_token')
        batch_token = request.form.get('batch_token')
        changes_json = request.form.get('changes')
        
        if not all([original_token, batch_token, changes_json]):
            return {"success": False, "error": "Missing required parameters"}, 400
        
        # Parse changes
        try:
            changes = json.loads(changes_json)
        except:
            return {"success": False, "error": "Invalid changes data"}, 400
        
        # Get original file data from RESULT_STORE
        original_entry = RESULT_STORE.get(original_token)
        if not original_entry:
            return {"success": False, "error": "Original file not found"}, 404
        
        # Parse original file data - use stored original order data if available
        try:
            if "original_order_json" in original_entry:
                # Use stored original order data
                original_rows = eval(original_entry.get("original_order_json", b"[]"))
            else:
                # Fallback: reconstruct from processed results
                processed_rows = eval(original_entry.get("rows_json", b"[]"))
                original_rows = []
                for row in processed_rows:
                    order_row = {
                        'fish name': row.get('fish name', ''),
                        'packed size': row.get('packed size', ''),
                        'total carton': row.get('order_carton', 0),
                        'weight_mc': '',
                        'order_kg_per_ctn': row.get('order_kg_per_ctn', 0)
                    }
                    original_rows.append(order_row)
        except:
            return {"success": False, "error": "Could not parse original file data"}, 500
        
        # Apply changes to create updated order data
        updated_order_data = apply_changes_to_order_data(original_rows, changes)
        
        # Convert updated order data to proper format for compute_matches
        updated_order_rows = []
        for row in updated_order_data:
            # Convert back to original order format
            order_row = {
                'fish name': row.get('fish name', ''),
                'packed size': row.get('packed size', ''),
                'total carton': row.get('total carton', 0),
                'weight_mc': row.get('weight_mc', ''),
                'remark': row.get('remark', '')
            }
            updated_order_rows.append(order_row)
        
        # Get stock data for reprocessing
        stock_name = (original_entry.get("stock_name") or b"").decode("utf-8", errors="ignore")
        
        # Find stock file from the batch to get stock data
        session_id = get_session_id_by_batch_token(batch_token)
        if not session_id:
            return {"success": False, "error": "Batch session not found"}, 404
        
        # Get all stock file data (regular, import, extra-load)
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Load all stock sources
        combined_stock_rows = []

        # Load regular stock first (required)
        cursor.execute("SELECT stored_filename FROM uploaded_files WHERE session_id = ? AND file_type = 'stock' LIMIT 1", (session_id,))
        stock_file_row = cursor.fetchone()
        
        if stock_file_row:
            stock_file_path = os.path.join(UPLOAD_FOLDER, stock_file_row['stored_filename'])
            try:
                with open(stock_file_path, 'rb') as f:
                    file_like = io.BytesIO(f.read())
                    file_like.filename = stock_file_path
                    stock_ds = load_tabular(file_like)
                    combined_stock_rows.extend(stock_ds.rows)
            except Exception as e:
                print(f"Error loading regular stock file: {e}")

        # Load import stock (optional)
        cursor.execute("SELECT stored_filename FROM uploaded_files WHERE session_id = ? AND file_type = 'import_stock'", (session_id,))
        import_stock_files = cursor.fetchall()

        for import_file_row in import_stock_files:
            import_stock_path = os.path.join(UPLOAD_FOLDER, import_file_row['stored_filename'])
            try:
                with open(import_stock_path, 'rb') as f:
                    file_like = io.BytesIO(f.read())
                    file_like.filename = import_stock_path
                    import_stock_ds = load_tabular(file_like)
                    combined_stock_rows.extend(import_stock_ds.rows)
            except Exception as e:
                print(f"Error loading import stock file {import_file_row['stored_filename']}: {e}")

        # Load extra-load stock (optional)
        cursor.execute("SELECT stored_filename FROM uploaded_files WHERE session_id = ? AND file_type = 'extra_load_stock'", (session_id,))
        extra_load_stock_files = cursor.fetchall()

        for extra_file_row in extra_load_stock_files:
            extra_stock_path = os.path.join(UPLOAD_FOLDER, extra_file_row['stored_filename'])
            try:
                with open(extra_stock_path, 'rb') as f:
                    file_like = io.BytesIO(f.read())
                    file_like.filename = extra_stock_path
                    extra_stock_ds = load_tabular(file_like)
                    combined_stock_rows.extend(extra_stock_ds.rows)
            except Exception as e:
                print(f"Error loading extra-load stock file {extra_file_row['stored_filename']}: {e}")

        conn.close()

        if not combined_stock_rows:
            return {"success": False, "error": "No stock files found"}, 404

        # Create combined stock dataset
        from dataclasses import dataclass
        @dataclass
        class CombinedDataset:
            rows: list
            name: str
            sheet_names: list

        stock_ds = CombinedDataset(
            rows=combined_stock_rows,
            name="combined_stock",
            sheet_names=["Combined Stock"]
        )
        
        # Recompute matches with updated order data
        new_result_rows = compute_matches(stock_ds.rows, updated_order_rows)
        
        # Calculate new summary
        def sum_required_kg(rows):
            return round(sum(float(r.get('required_kg', 0) or 0) for r in rows), 3)

        def sum_ready_kg(rows):
            return round(sum(float(r.get('stock_carton', 0) or 0) * float(r.get('stock_kg_per_ctn', 0) or 0) for r in rows), 3)

        def sum_fulfillable_kg(rows):
            return round(sum(float(r.get('can_fulfill_carton', 0) or 0) * float(r.get('order_kg_per_ctn', 0) or 0) for r in rows), 3)

        new_summary = {
            "total_items": len(new_result_rows),
            "full": sum(1 for r in new_result_rows if r.get('status') == 'Full'),
            "not_full": sum(1 for r in new_result_rows if r.get('status') == 'Not Full'),
            "not_have": sum(1 for r in new_result_rows if r.get('status') == 'Not have'),
            "total_kg_all": sum_required_kg(new_result_rows),
            "total_kg_full": sum_required_kg([r for r in new_result_rows if r.get('status') == 'Full']),
            "total_kg_not_full": sum_required_kg([r for r in new_result_rows if r.get('status') == 'Not Full']),
            "total_kg_not_have": sum_required_kg([r for r in new_result_rows if r.get('status') == 'Not have']),
            "ready_kg": sum_ready_kg(new_result_rows),
            "fulfillable_kg": sum_fulfillable_kg(new_result_rows)
        }
        
        # Create new Excel and PDF files
        excel_bytes = rows_to_excel_bytes(new_result_rows)
        pdf_bytes = rows_to_pdf_bytes(new_result_rows)
        
        # Update RESULT_STORE
        order_name = (original_entry.get("order_name") or b"").decode("utf-8", errors="ignore")
        order_basename = os.path.splitext(order_name)[0] if order_name else "order"
        
        RESULT_STORE[original_token] = {
            "excel": excel_bytes,
            "pdf": pdf_bytes,
            "excel_name": f"{order_basename} Calculation.xlsx",
            "pdf_name": f"{order_basename} Calculation.pdf",
            "rows_json": io.BytesIO(str(new_result_rows).encode("utf-8")).getvalue(),
            "summary_json": io.BytesIO(str(new_summary).encode("utf-8")).getvalue(),
            "stock_name": original_entry.get("stock_name", b""),
            "order_name": original_entry.get("order_name", b""),
            "original_order_json": io.BytesIO(str(updated_order_data).encode("utf-8")).getvalue(),
        }
        
        # Update database if session exists
        try:
            if session_id:
                # Find order file ID
                conn = sqlite3.connect(DATABASE_PATH)
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                cursor.execute("SELECT id, stored_filename FROM uploaded_files WHERE session_id=? AND file_type='order' AND original_filename=? ORDER BY id DESC LIMIT 1", 
                             (session_id, order_name))
                order_file_record = cursor.fetchone()
                
                if order_file_record:
                    order_file_id = order_file_record[0]
                    stored_filename = order_file_record[1]
                    
                    print(f"Applying changes for session {session_id}, order file {order_file_id}")
                    
                    # Save the updated order data as a new file version
                    try:
                        # Create a new Excel file with the updated order data
                        from openpyxl import Workbook
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Order"
                        
                        # Add headers
                        headers = ['Fish Name', 'Packed Size', 'Pack', 'Total Carton', 'Weight MC', 'Order KG/CTN', 'Remark']
                        for col, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col, value=header)
                        
                        # Add data
                        for row_idx, order_row in enumerate(updated_order_data, 2):
                            weight_mc = order_row.get('weight_mc', '')
                            order_kg_per_ctn = order_row.get('order_kg_per_ctn', 0)
                            
                            # If weight_mc is empty but we have order_kg_per_ctn, use it
                            if not weight_mc and order_kg_per_ctn:
                                weight_mc = str(order_kg_per_ctn)
                                
                            ws.cell(row=row_idx, column=1, value=order_row.get('fish name', ''))
                            ws.cell(row=row_idx, column=2, value=order_row.get('packed size', ''))
                            ws.cell(row=row_idx, column=3, value=order_row.get('pack', ''))
                            ws.cell(row=row_idx, column=4, value=order_row.get('total carton', 0))
                            ws.cell(row=row_idx, column=5, value=weight_mc)
                            ws.cell(row=row_idx, column=6, value=order_kg_per_ctn)
                            ws.cell(row=row_idx, column=7, value=order_row.get('remark', ''))
                        
                        # Save to the same filename (overwrite)
                        updated_file_path = os.path.join(UPLOAD_FOLDER, stored_filename)
                        wb.save(updated_file_path)
                        
                        # Also save as a revision in the files system
                        try:
                            revision_note = f"Editor changes applied on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
                            save_revised_file(session_id, order_name, updated_file_path, revision_note)
                        except Exception as rev_e:
                            print(f"Warning: Failed to save revision file: {rev_e}")
                        
                    except Exception as e:
                        print(f"Warning: Failed to save updated order file: {e}")
                    
                    # Delete old results
                    deleted_count = cursor.execute('DELETE FROM processing_results WHERE session_id=? AND order_file_id=?', (session_id, order_file_id)).rowcount
                    print(f"Deleted {deleted_count} old processing results")
                    
                    # Commit the deletion before saving new results
                    conn.commit()
                    
                    # Save new results (this function handles its own connection)
                    save_processing_results(session_id, new_result_rows, order_file_id)
                    print(f"Saved {len(new_result_rows)} new processing results")
                
                # Update batch aggregated summary in processing_sessions table
                # Get all order tokens for this batch to recalculate totals
                tokens = BATCH_STORE.get(batch_token, [])
                agg_total_items = 0
                agg_full = 0
                agg_nf = 0
                agg_nh = 0
                agg_kg_all = 0
                agg_kg_full = 0
                agg_kg_nf = 0
                agg_kg_nh = 0
                
                for token in tokens:
                    entry = RESULT_STORE.get(token, {})
                    try:
                        summary = eval(entry.get("summary_json", b"{}"))
                        agg_total_items += summary.get("total_items", 0)
                        agg_full += summary.get("full", 0)
                        agg_nf += summary.get("not_full", 0)
                        agg_nh += summary.get("not_have", 0)
                        agg_kg_all += summary.get("total_kg_all", 0)
                        agg_kg_full += summary.get("total_kg_full", 0)
                        agg_kg_nf += summary.get("total_kg_not_full", 0)
                        agg_kg_nh += summary.get("total_kg_not_have", 0)
                    except:
                        pass
                
                cursor.execute(
                    '''UPDATE processing_sessions
                       SET total_items=?, full_items=?, not_full_items=?, not_have_items=?,
                           total_kg=?, full_kg=?, not_full_kg=?, not_have_kg=?
                     WHERE id=?''',
                    (
                        int(agg_total_items), int(agg_full), int(agg_nf), int(agg_nh),
                        float(agg_kg_all), float(agg_kg_full), float(agg_kg_nf), float(agg_kg_nh),
                        session_id,
                    )
                )
                print(f"Updated session {session_id} summary: items={agg_total_items}, full={agg_full}, not_full={agg_nf}, not_have={agg_nh}")
                
                conn.commit()
                conn.close()
        except Exception as e:
            print(f"Warning: Could not update database: {e}")
        
        # Save comparison history to database only when changes are applied
        try:
            if session_id:
                # Find order file ID
                conn = sqlite3.connect(DATABASE_PATH)
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM uploaded_files WHERE session_id=? AND file_type='order' AND original_filename=? ORDER BY id DESC LIMIT 1", 
                             (session_id, order_name))
                order_file_record = cursor.fetchone()
                
                if order_file_record:
                    order_file_id = order_file_record[0]
                    
                    # Get the comparison data that was used to generate these changes
                    # We need to reconstruct it from the changes
                    comparison_data = {
                        'summary': {
                            'added': len([c for c in changes if c.get('type') == 'added']),
                            'modified': len([c for c in changes if c.get('type') == 'modified']),
                            'deleted': len([c for c in changes if c.get('type') == 'deleted']),
                            'unchanged': 0  # We don't track unchanged in apply changes
                        },
                        'changes': changes
                    }
                    
                    save_file_comparison_history(
                        session_id=session_id,
                        order_file_id=order_file_id,
                        original_token=original_token,
                        batch_token=batch_token,
                        comparison_data=comparison_data,
                        changes_applied=changes
                    )
                
                conn.close()
        except Exception as e:
            print(f"Warning: Failed to save comparison history: {e}")
        
        return {"success": True}
        
    except Exception as e:
        return {"success": False, "error": str(e)}, 500


@app.get("/get-batch-files/<batch_token>")
def get_batch_files(batch_token: str):
    """Get all files (stock and order) for a batch with revision tracking."""
    try:
        session_id = get_session_id_by_batch_token(batch_token)
        if not session_id:
            return {"success": False, "error": "Session not found"}, 404
        
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Get stock files (should be original uploads only)
        cursor.execute('''
            SELECT * FROM uploaded_files 
            WHERE session_id = ? AND file_type = 'stock'
            ORDER BY uploaded_at ASC
        ''', (session_id,))
        stock_files = [dict(row) for row in cursor.fetchall()]
        
        # Get original order files (is_revision = 0 or NULL)
        cursor.execute('''
            SELECT * FROM uploaded_files 
            WHERE session_id = ? AND file_type = 'order' AND (is_revision = 0 OR is_revision IS NULL)
            ORDER BY uploaded_at ASC
        ''', (session_id,))
        original_order_files = [dict(row) for row in cursor.fetchall()]
        
        # Get revised order files (is_revision = 1)
        cursor.execute('''
            SELECT * FROM uploaded_files 
            WHERE session_id = ? AND file_type = 'order' AND is_revision = 1
            ORDER BY uploaded_at DESC
        ''', (session_id,))
        revised_order_files = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        return {
            "success": True,
            "stock_files": stock_files,
            "original_order_files": original_order_files,
            "revised_order_files": revised_order_files
        }
        
    except Exception as e:
        return {"success": False, "error": str(e)}, 500


@app.get("/download-file/<int:file_id>")
def download_file(file_id: int):
    """Download a file by its ID."""
    try:
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Get file info
        cursor.execute('SELECT * FROM uploaded_files WHERE id = ?', (file_id,))
        file_record = cursor.fetchone()
        
        if not file_record:
            conn.close()
            flash("File not found.", "error")
            return redirect(url_for("index"))
        
        file_record = dict(file_record)
        conn.close()
        
        # Build file path
        file_path = os.path.join(UPLOAD_FOLDER, file_record['stored_filename'])
        
        # Check if file exists on disk
        if not os.path.exists(file_path):
            flash("File not found on disk.", "error")
            return redirect(url_for("index"))
        
        # Send file
        return send_file(
            file_path,
            as_attachment=True,
            download_name=file_record['original_filename']
        )
        
    except Exception as e:
        flash(f"Error downloading file: {str(e)}", "error")
        return redirect(url_for("index"))


def apply_changes_to_order_data(original_rows, changes):
    """Apply comparison changes to create updated order data."""
    # Start with all original data (keep unchanged items)
    updated_rows = [row.copy() for row in original_rows]
    
    # Create a lookup for original rows for quick access
    original_lookup = {}
    for i, row in enumerate(original_rows):
        key = (
            canonicalize_product(row.get('fish name', '')),
            canonicalize_pack(row.get('packed size', ''))
        )
        original_lookup[key] = i
    
    # Apply changes
    for change in changes:
        key = (
            canonicalize_product(change.get('fish_name', '')),
            canonicalize_pack(change.get('packed_size', ''))
        )
        
        if change['status'] == 'added':
            # Add new item to the list
            weight_mc = ''
            order_kg_per_ctn = 0
            try:
                # Try to parse kg from packed size if weight_mc is empty
                order_kg_per_ctn = parse_kg_per_carton(change['packed_size'])
            except:
                pass
            
            new_row = {
                'fish name': change['fish_name'],
                'packed size': change['packed_size'],
                'pack': change.get('new_pack', ''),
                'total carton': change['new_quantity'],
                'weight_mc': weight_mc,
                'order_kg_per_ctn': order_kg_per_ctn,
                'remark': change.get('new_remark', '')
            }
            updated_rows.append(new_row)
            
        elif change['status'] == 'modified':
            # Update existing item quantity and weight_mc
            if key in original_lookup:
                row_index = original_lookup[key]
                updated_rows[row_index]['total carton'] = change['new_quantity']
                if 'new_weight_mc' in change:
                    updated_rows[row_index]['weight_mc'] = change['new_weight_mc']
                if 'new_order_kg_per_ctn' in change:
                    updated_rows[row_index]['order_kg_per_ctn'] = change['new_order_kg_per_ctn']
                if 'new_pack' in change:
                    updated_rows[row_index]['pack'] = change['new_pack']
                if 'new_remark' in change:
                    updated_rows[row_index]['remark'] = change['new_remark']
                
        elif change['status'] == 'deleted':
            # Remove item from the list
            if key in original_lookup:
                row_index = original_lookup[key]
                # Mark for removal (we'll filter out later to maintain indices)
                updated_rows[row_index] = None
    
    # Filter out deleted items (marked as None)
    updated_rows = [row for row in updated_rows if row is not None]
    
    return updated_rows


@app.route('/save-finished-order', methods=['POST'])
def save_finished_order_route():
    try:
        data = request.get_json()
        batch_token = data.get('batch_token')
        order_token = data.get('order_token')
        order_name = data.get('order_name')
        weight = data.get('weight')
        action = data.get('action', 'save')
        
        print(f"Received request to {action} finished order:")
        print(f"  batch_token: {batch_token}")
        print(f"  order_token: {order_token}")
        print(f"  order_name: {order_name}")
        print(f"  weight: {weight}")
        
        # Debug: Check if the table exists
        conn_debug = sqlite3.connect(DATABASE_PATH)
        cursor_debug = conn_debug.cursor()
        cursor_debug.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='finished_orders'")
        table_exists = cursor_debug.fetchone()
        print(f"finished_orders table exists: {table_exists is not None}")
        conn_debug.close()
        
        if not all([batch_token, order_token, order_name]):
            return {"success": False, "error": "Missing required data"}
        
        # Get session by batch token
        session_id = get_session_id_by_batch_token(batch_token)
        print(f"Found session_id: {session_id}")
        if not session_id:
            return {"success": False, "error": "Session not found"}
        
        conn = sqlite3.connect(DATABASE_PATH)
        cursor = conn.cursor()
        
        if action == 'save':
            # Save finished order
            print(f"Inserting finished order into database...")
            try:
                cursor.execute('''
                    INSERT OR REPLACE INTO finished_orders 
                    (session_id, batch_token, order_token, order_name, weight)
                    VALUES (?, ?, ?, ?, ?)
                ''', (session_id, batch_token, order_token, order_name, weight))
                print(f"Inserted successfully, affected rows: {cursor.rowcount}")
                
                # Verify the insert
                cursor.execute('SELECT * FROM finished_orders WHERE session_id = ? AND order_token = ?', 
                              (session_id, order_token))
                result = cursor.fetchone()
                print(f"Verification query result: {result}")
                
            except Exception as e:
                print(f"Error inserting finished order: {e}")
                conn.close()
                return {"success": False, "error": f"Database insert error: {str(e)}"}
                
        elif action == 'remove':
            # Remove finished order
            print(f"Removing finished order from database...")
            try:
                cursor.execute('''
                    DELETE FROM finished_orders 
                    WHERE session_id = ? AND order_token = ?
                ''', (session_id, order_token))
                print(f"Removed successfully, affected rows: {cursor.rowcount}")
            except Exception as e:
                print(f"Error removing finished order: {e}")
                conn.close()
                return {"success": False, "error": f"Database delete error: {str(e)}"}
        
        conn.commit()
        conn.close()
        
        return {"success": True}
        
    except Exception as e:
        print(f"Error in save_finished_order: {e}")
        return {"success": False, "error": str(e)}, 500

@app.route('/get-finished-orders/<batch_token>')
def get_finished_orders(batch_token):
    try:
        print(f"Getting finished orders for batch: {batch_token}")
        
        # Get session by batch token
        session_id = get_session_id_by_batch_token(batch_token)
        if not session_id:
            return {"success": False, "error": "Session not found"}
        
        print(f"Found session_id: {session_id}")
        
        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT order_token, order_name, weight, finished_at
            FROM finished_orders
            WHERE session_id = ?
            ORDER BY finished_at DESC
        ''', (session_id,))
        
        finished_orders = []
        for row in cursor.fetchall():
            finished_orders.append({
                'token': row['order_token'],
                'name': row['order_name'],
                'weight': row['weight'],
                'finished_at': row['finished_at']
            })
        
        conn.close()
        
        print(f"Found {len(finished_orders)} finished orders")
        return {"success": True, "finished_orders": finished_orders}
        
    except Exception as e:
        print(f"Error getting finished orders: {e}")
        return {"success": False, "error": str(e)}, 500


@app.route('/api/batch/<int:session_id>/results')
def get_batch_results(session_id):
    """API endpoint to get updated processing results for a batch."""
    try:
        # Check if user is logged in to either module
        if not (session.get('packing_logged_in') or session.get('raw_materials_logged_in')):
            return {"success": False, "error": "Not logged in"}, 401

        conn = sqlite3.connect(DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Check if session exists
        cursor.execute('SELECT id FROM processing_sessions WHERE id = ?', (session_id,))
        if not cursor.fetchone():
            conn.close()
            return {"success": False, "error": "Batch not found"}, 404

        # Get processing results with order file information, ordered by most recent first
        cursor.execute('''
            SELECT pr.*,
                   uf.original_filename as order_filename,
                   uf.uploaded_at as file_uploaded_at
            FROM processing_results pr
            LEFT JOIN uploaded_files uf ON pr.order_file_id = uf.id
            WHERE pr.session_id = ?
            ORDER BY pr.id DESC, pr.fish_name ASC
        ''', (session_id,))
        results = [dict(row) for row in cursor.fetchall()]

        # Debug information
        print(f"API: Found {len(results)} results for session {session_id}")
        if results:
            print(f"API: Latest result ID: {results[0].get('id')}, Fish: {results[0].get('fish_name')}")

        # Get updated session summary
        cursor.execute('''
            SELECT total_items, full_items, not_full_items, not_have_items
            FROM processing_sessions
            WHERE id = ?
        ''', (session_id,))
        session_data = cursor.fetchone()
        
        # Also get the latest processing result timestamp for cache busting
        cursor.execute('''
            SELECT MAX(strftime('%s', pr.created_at)) as latest_result_time
            FROM processing_results pr
            WHERE pr.session_id = ?
        ''', (session_id,))
        latest_result = cursor.fetchone()
        
        conn.close()

        session_summary = dict(session_data) if session_data else {}
        
        # Use the latest timestamp available for cache busting
        latest_timestamp = int(time.time())
        if latest_result and latest_result['latest_result_time']:
            latest_timestamp = max(latest_timestamp, int(latest_result['latest_result_time']))

        return {
            "success": True, 
            "results": results,
            "summary": session_summary,
            "timestamp": latest_timestamp,
            "debug": {
                "result_count": len(results),
                "session_id": session_id
            }
        }
        
    except Exception as e:
        print(f"Error getting batch results: {e}")
        return {"success": False, "error": str(e)}, 500

if __name__ == "__main__":
    app.run(debug=True)


